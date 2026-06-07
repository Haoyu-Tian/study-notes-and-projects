# -*- coding: utf-8 -*-
"""
人脸识别智能考勤管理系统 —— SQLite 版（3.0）
=============================================
在优化版（2.0）基础上，将 JSON 文件存储全部替换为 SQLite 数据库。

核心变化：
    1. 存储层重构
       去掉 attendance_state.json / attendance_log.json 两个文件，
       统一使用单一 SQLite 数据库文件（attendance.db）管理所有数据。

    2. 异步写入器升级
       AsyncWriter（JSON 异步写）→ DBWriter（SQLite 异步串行写）。
       所有写操作经由单一写线程串行执行，彻底消除多线程 SQLite 竞争。

    3. 历史数据自动保留
       原版 JSON 按日期覆盖，只保留当天数据。
       SQLite 版通过 date 字段区分每天，历史数据永久保留，
       支持按日期查询任意历史考勤记录。

    4. 导出数据源升级
       原版导出时从内存快照读取（存在并发风险）。
       SQLite 版导出时直接从数据库读取（保证数据最新且线程安全）。

    5. 对外接口完全不变
       AttendanceManager 的所有公共方法签名与 2.0 版本完全一致，
       上层代码（UIRenderer、AttendanceSystem）无需任何修改。

数据库结构（attendance.db）：
    ┌─ attendance_status ──────────────────────────────────────┐
    │  name TEXT PK │ status TEXT │ last_time REAL │ date TEXT │
    └──────────────────────────────────────────────────────────┘
    ┌─ attendance_log ─────────────────────────────────────────┐
    │  id INTEGER PK │ name TEXT │ type TEXT │ time TEXT        │
    │  date TEXT     │ (name+date 联合索引，date 单独索引)       │
    └──────────────────────────────────────────────────────────┘
"""

import cv2
import numpy as np
import os
import sqlite3
import time
import logging
import logging.handlers
from datetime import datetime, timedelta
from insightface.app import FaceAnalysis
from PIL import Image, ImageDraw, ImageFont
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import threading
import queue
from collections import deque, OrderedDict


# ============================================================
#                        日志配置模块
# ============================================================

def setup_logger(log_dir: str = "logs") -> logging.Logger:
    """
    初始化全局日志记录器。

    配置策略：
        - 控制台 Handler：输出 INFO 及以上级别，适合运行时监控。
        - 文件 Handler  ：输出 DEBUG 及以上级别，按天滚动，保留 7 天。
        - 统一格式      ：[时间戳] 级别 模块名 - 消息内容。

    Args:
        log_dir (str): 日志文件存放目录，默认 "logs"，不存在时自动创建。

    Returns:
        logging.Logger: 配置完成的日志记录器实例，名称为 "attendance"。
    """
    os.makedirs(log_dir, exist_ok=True)
    logger = logging.getLogger("attendance")
    logger.setLevel(logging.DEBUG)
    fmt = logging.Formatter(
        "[%(asctime)s] %(levelname)-8s %(name)s - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"
    )
    # 控制台 Handler：INFO 及以上，适合日常运行观察
    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)
    # 滚动文件 Handler：DEBUG 及以上，每天零点生成新文件，保留最近 7 天
    fh = logging.handlers.TimedRotatingFileHandler(
        os.path.join(log_dir, "attendance.log"),
        when="midnight", backupCount=7, encoding="utf-8"
    )
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)
    logger.addHandler(ch)
    logger.addHandler(fh)
    return logger


# 模块级全局日志记录器，供所有模块统一使用
logger = setup_logger()


# ============================================================
#                         全局配置
# ============================================================

# ---------- 文件路径 ----------
KNOWN_FACES_DIR = "known_faces"       # 已知人员照片目录
DB_FILE         = "attendance.db"     # SQLite 数据库文件（替换原来两个 JSON 文件）
EXPORT_DIR      = "attendance_reports"  # Excel 报表输出目录

# ---------- 识别与打卡参数 ----------
SIMILARITY_THRESHOLD = 0.35  # 人脸相似度阈值，低于此值视为未知人员
CONFIRM_FRAMES       = 6     # 连续检测到同一人至少 N 帧后才触发打卡
COOLDOWN_SECONDS     = 5     # 同一人两次打卡之间的最短冷却时间（秒）
DETECT_INTERVAL      = 2     # 每隔 N 帧执行一次人脸检测，降低 CPU 占用

# ---------- 摄像头分辨率 ----------
CAMERA_WIDTH  = 1280
CAMERA_HEIGHT = 720

# ---------- UI 字体路径 ----------
FONT_PATH      = "C:/Windows/Fonts/msyh.ttc"    # 微软雅黑常规体
FONT_PATH_BOLD = "C:/Windows/Fonts/msyhbd.ttc"  # 微软雅黑粗体

# ---------- 窗口与布局 ----------
WINDOW_NAME  = "Face Attendance System"  # OpenCV 窗口标题
PANEL_RATIO  = 0.22                      # 右侧信息面板占总窗口宽度的比例

# ---------- 通知与缓存 ----------
MAX_NOTIFICATIONS = 4    # 同一时刻最多并排显示的打卡通知条数
PANEL_STATUS_TTL  = 0.5  # 面板统计层缓存有效期（秒）
FONT_CACHE_MAX    = 32   # LRU 字体缓存最大条目数

# ---------- 颜色常量（BGR 格式）----------
C_BG_PANEL = (45,  45,  45 )  # 面板背景色（深灰）
C_ACCENT   = (0,   200, 100)  # 主强调色（绿色）
C_ACCENT2  = (0,   160, 255)  # 副强调色（蓝色）
C_WHITE    = (240, 240, 240)  # 近白色
C_GRAY     = (160, 160, 160)  # 灰色，用于次要文字
C_GREEN    = (0,   220, 80 )  # 绿色，表示签到/在岗
C_CYAN     = (200, 220, 0  )  # 青黄色，用于工时信息
C_ORANGE   = (0,   165, 255)  # 橙色，表示签退/警告
C_RED      = (60,  60,  220)  # 红色（BGR），表示错误/异常
C_BORDER   = (80,  80,  80 )  # 边框线颜色
C_UNKNOWN  = (0,   220, 220)  # 黄色，表示未知人员人脸框


# ============================================================
#                  字体缓存（LRU + 容量上限）
# ============================================================

class _LRUFontCache:
    """
    基于 OrderedDict 实现的 LRU（最近最少使用）字体缓存。

    解决问题：
        普通字典缓存无上限，长期运行时字体对象无限积累造成内存泄漏。

    LRU 淘汰策略：
        - 命中缓存时：将该条目移到末尾（标记为最近使用）。
        - 新增条目时：若已达上限，弹出头部（最久未使用）的条目。

    线程安全：
        内部使用 threading.Lock 保护，支持多线程并发访问。

    字体加载降级策略（三级）：
        粗体路径失败 → 回退常规体路径 → 回退 PIL 内置字体，
        每一级失败都记录日志，不静默吞异常。
    """

    def __init__(self, maxsize: int = FONT_CACHE_MAX):
        """
        Args:
            maxsize (int): 缓存最大条目数，默认 FONT_CACHE_MAX（32）。
        """
        self._cache: OrderedDict = OrderedDict()
        self._maxsize = maxsize
        self._lock = threading.Lock()

    def get(self, size: int, bold: bool) -> ImageFont.FreeTypeFont:
        """
        获取指定字号和字重的 PIL 字体对象（优先从缓存取，未命中则加载）。

        Args:
            size (int) : 字号（像素）。
            bold (bool): 是否使用粗体。

        Returns:
            ImageFont.FreeTypeFont: PIL 字体对象。
        """
        key = (size, bold)
        with self._lock:
            if key in self._cache:
                # 命中缓存：移到末尾，标记为最近使用
                self._cache.move_to_end(key)
                return self._cache[key]
            # 未命中：尝试加载字体文件，三级降级策略
            try:
                path = FONT_PATH_BOLD if bold else FONT_PATH
                font = ImageFont.truetype(path, size)
            except OSError as e:
                logger.warning("字体加载失败(%s size=%d)，回退: %s", path, size, e)
                try:
                    # 第一级回退：使用常规体路径
                    font = ImageFont.truetype(FONT_PATH, size)
                except OSError:
                    # 第二级回退：使用 PIL 内置位图字体（不支持中文，仅保证不崩溃）
                    logger.error("默认字体也无法加载，使用 PIL 内置字体")
                    font = ImageFont.load_default()
            # 超出容量：淘汰最久未使用的条目（头部）
            if len(self._cache) >= self._maxsize:
                evicted = self._cache.popitem(last=False)
                logger.debug("字体缓存淘汰: key=%s", evicted[0])
            self._cache[key] = font
            return font


# 模块级 LRU 字体缓存单例
_font_cache = _LRUFontCache()


def get_font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont:
    """
    获取 PIL 字体对象的公共接口（委托给 LRU 缓存）。

    Args:
        size (int) : 字号（像素）。
        bold (bool): 是否使用粗体，默认 False。

    Returns:
        ImageFont.FreeTypeFont: 对应的 PIL 字体对象。
    """
    return _font_cache.get(size, bold)


# ============================================================
#                       文字渲染工具
# ============================================================

def draw_texts(img_bgr: np.ndarray, items: list) -> np.ndarray:
    """
    批量在 BGR 图像上渲染中文文字（使用 PIL 绘制，避免 OpenCV 中文乱码）。

    性能优化：整个批次只做一次 BGR→RGB→BGR 颜色空间转换，
    相比逐条调用可显著减少转换开销。

    Args:
        img_bgr (np.ndarray): 输入的 BGR 图像。
        items   (list)       : 文字绘制参数列表，每项格式为：
            (text, (x, y), color_bgr, size, bold)

    Returns:
        np.ndarray: 绘制文字后的 BGR 图像。
    """
    if not items:
        return img_bgr
    # BGR → RGB，转为 PIL 可处理的格式
    img_rgb = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2RGB)
    pil_img = Image.fromarray(img_rgb)
    draw    = ImageDraw.Draw(pil_img)
    for text, pos, color_bgr, size, bold in items:
        font      = get_font(size, bold)
        # PIL 使用 RGB 颜色顺序，将 BGR 分量反转
        pil_color = (color_bgr[2], color_bgr[1], color_bgr[0])
        draw.text(pos, text, font=font, fill=pil_color)
    # RGB → BGR，转回 OpenCV 格式
    return cv2.cvtColor(np.array(pil_img), cv2.COLOR_RGB2BGR)


def blend_rect(img, pt1, pt2, color, alpha=0.55):
    """
    在图像指定区域绘制半透明填充矩形（Alpha 混合）。

    通过 cv2.addWeighted 将纯色矩形与原图 ROI 混合，实现毛玻璃/遮罩效果。
    坐标会自动裁剪到图像边界内，防止越界访问。

    Args:
        img   (np.ndarray): 目标 BGR 图像（原地修改）。
        pt1   (tuple)     : 矩形左上角坐标 (x1, y1)。
        pt2   (tuple)     : 矩形右下角坐标 (x2, y2)。
        color (tuple)     : 填充颜色，BGR 格式。
        alpha (float)     : 遮罩不透明度（0.0=完全透明，1.0=完全不透明），默认 0.55。

    Returns:
        np.ndarray: 混合后的 BGR 图像（与输入共享内存）。
    """
    x1, y1 = pt1; x2, y2 = pt2
    # 坐标裁剪，防止超出图像边界
    x1 = max(0, x1); y1 = max(0, y1)
    x2 = min(img.shape[1] - 1, x2)
    y2 = min(img.shape[0] - 1, y2)
    if x2 <= x1 or y2 <= y1:
        return img  # 无效区域，直接返回
    roi = img[y1:y2, x1:x2]
    cv2.addWeighted(np.full_like(roi, color), alpha, roi, 1 - alpha, 0, roi)
    img[y1:y2, x1:x2] = roi
    return img


# ============================================================
#                      考勤计算模块（纯函数）
# ============================================================

def calc_work_sessions(records: list):
    """
    根据原始打卡流水计算工作段、总工时及异常情况（纯函数，不依赖任何外部状态）。

    算法（栈式 FIFO 配对）：
        - 遇到签到：压入栈。
        - 遇到签退：弹出最早的签到与之配对，计算本段工时。
        - 签退时栈为空：标记"无对应签到"异常。
        - 处理完毕栈非空：标记"未签退"异常。

    兼容性说明：
        records 中每项格式为 {"type": "签到"/"签退", "time": "YYYY-MM-DD HH:MM:SS"}，
        与 JSON 版和 SQLite 版的查询结果格式完全一致，无需适配。

    Args:
        records (list): 打卡记录列表，按时间升序排列。

    Returns:
        tuple:
            - sessions  (list)     : 工作段列表，每项含 sign_in、sign_out、
                                     duration（timedelta）、note 字段。
            - total_dur (timedelta): 所有有效工作段的累计工时。
            - anomalies (list)     : 异常描述字符串列表。
    """
    sessions, anomalies, stack_in = [], [], []
    for rec in records:
        # 兼容完整 datetime 字符串（取后8位时间部分）和纯时间字符串
        t_str = rec["time"][11:] if len(rec["time"]) > 8 else rec["time"]
        if rec["type"] == "签到":
            stack_in.append(t_str)
        else:
            if stack_in:
                # FIFO：取最早的签到与当前签退配对
                sign_in = stack_in.pop(0)
                dur     = _tdiff(sign_in, t_str)
                note    = "" if dur.total_seconds() >= 0 else "时间异常"
                sessions.append({"sign_in": sign_in, "sign_out": t_str,
                                  "duration": dur, "note": note})
                if dur.total_seconds() < 0:
                    anomalies.append(f"签退早于签到:{sign_in}→{t_str}")
            else:
                # 签退时栈为空，无对应签到
                sessions.append({"sign_in": None, "sign_out": t_str,
                                  "duration": None, "note": "无对应签到"})
                anomalies.append(f"无对应签到:{t_str}")
    # 处理栈中剩余的未配对签到（即未签退）
    for si in stack_in:
        sessions.append({"sign_in": si, "sign_out": None,
                          "duration": None, "note": "未签退"})
        anomalies.append(f"未签退:{si}")
    # 按签到时间升序排列，None 排到末尾
    sessions.sort(key=lambda s: s["sign_in"] or "99:99:99")
    # 累计所有有效（正数工时）工作段
    total = sum(
        (s["duration"] for s in sessions
         if s["duration"] and s["duration"].total_seconds() > 0),
        timedelta()
    )
    return sessions, total, anomalies


def _tdiff(t1: str, t2: str) -> timedelta:
    """
    计算两个时间字符串（HH:MM:SS 格式）的差值。

    Args:
        t1 (str): 起始时间字符串，格式 "%H:%M:%S"。
        t2 (str): 结束时间字符串，格式 "%H:%M:%S"。

    Returns:
        timedelta: t2 - t1 的时间差（可能为负数，表示时间异常）。
    """
    return datetime.strptime(t2, "%H:%M:%S") - datetime.strptime(t1, "%H:%M:%S")


def fmt_dur(td) -> str:
    """
    将 timedelta 对象格式化为人类可读的工时字符串。

    格式规则：
        - None 或无效值 → "--"
        - 负数（异常）  → "异常"
        - 有小时        → "XhYYm"
        - 仅有分钟      → "YmZZs"
        - 仅有秒        → "ZZs"

    Args:
        td (timedelta | None): 时间差对象。

    Returns:
        str: 格式化后的工时字符串。
    """
    if not td: return "--"
    s = int(td.total_seconds())
    if s < 0: return "异常"
    h, m, sec = s // 3600, (s % 3600) // 60, s % 60
    if h > 0:   return f"{h}h{m:02d}m"
    elif m > 0: return f"{m}m{sec:02d}s"
    return f"{sec}s"


# ============================================================
#                     多人帧计数器
# ============================================================

class MultiPersonCounter:
    """
    为每个识别到的人独立维护连续帧计数，满足阈值后触发打卡。

    Bug 修复：使用 ">=" 而非 "==" 判断，避免跳帧场景下永远无法触发的问题。
    触发后立即重置计数，避免同一驻留周期内重复触发打卡。
    """

    def __init__(self, confirm_frames: int):
        """
        Args:
            confirm_frames (int): 触发打卡所需的最少连续检测帧数。
        """
        self.confirm_frames = confirm_frames
        # 每个人名对应的当前连续检测帧数
        self._counters: dict = {}
        # 已触发打卡且尚未离开画面的人名集合（防止同一驻留周期重复触发）
        self._triggered: set = set()

    def update(self, detected_names: set) -> set:
        """
        用本帧检测到的人名集合更新内部计数，返回本帧应触发打卡的人名集合。

        Args:
            detected_names (set): 本帧识别成功的人名集合。

        Returns:
            set: 本帧需要执行打卡的人名集合（可能为空集）。
        """
        to_attend = set()
        # 消失的人：清零计数并解除触发锁
        for name in list(self._counters.keys()):
            if name not in detected_names:
                self._counters[name] = 0
                self._triggered.discard(name)
        # 本帧检测到的人：累加计数并判断是否达到触发条件
        for name in detected_names:
            self._counters[name] = self._counters.get(name, 0) + 1
            # 使用 >= 而非 == 避免跳帧导致永远无法触发
            if (self._counters[name] >= self.confirm_frames
                    and name not in self._triggered):
                to_attend.add(name)
                self._triggered.add(name)
                # 触发后重置，防止同一驻留周期内重复打卡
                self._counters[name] = 0
        return to_attend

    def reset_person(self, name: str):
        """
        手动重置指定人员的触发状态（通常在冷却期拒绝打卡后由外部调用）。

        Args:
            name (str): 需要重置状态的人员姓名。
        """
        self._triggered.discard(name)
        self._counters[name] = 0


# ============================================================
#                     多条通知队列
# ============================================================

class NotificationQueue:
    """
    支持多人同框时同时在视频区显示多条打卡通知。

    特性：
        - 每条通知独立计时，到期自动清除。
        - 同一人的旧通知会被新通知替换，避免重复堆叠。
        - 超出最大数量时移除最旧的通知（FIFO）。
        - 内部使用 threading.Lock 保证多线程安全。
    """

    def __init__(self, max_count: int = 4, duration: float = 3.0):
        """
        Args:
            max_count (int)  : 同时显示的最大通知条数，默认 4。
            duration  (float): 每条通知的显示持续时间（秒），默认 3.0。
        """
        self.max_count = max_count
        self.duration  = duration
        # 通知存储队列，每项包含 name/action/dur_str/expire 四个字段
        self._queue: deque = deque()
        self._lock = threading.Lock()

    def push(self, name: str, action: str, dur_str: str):
        """
        推入一条新的打卡通知。

        若同一人已有通知，先移除旧通知再推入新通知。
        若队列已满，移除最旧的通知腾出空间。

        Args:
            name    (str): 打卡人员姓名。
            action  (str): 打卡动作（"签到" 或 "签退"）。
            dur_str (str): 当前累计工时格式化字符串。
        """
        with self._lock:
            # 移除同一人的旧通知，防止重复堆叠
            self._queue = deque(n for n in self._queue if n["name"] != name)
            if len(self._queue) >= self.max_count:
                self._queue.popleft()  # 队列满时移除最旧的通知
            self._queue.append({
                "name": name, "action": action, "dur_str": dur_str,
                "expire": time.time() + self.duration  # 到期时间戳
            })

    def get_active(self) -> list:
        """
        获取当前所有未过期的通知列表，同时清理已过期的通知。

        Returns:
            list: 未过期通知的有序列表（按推入顺序，最旧的在前）。
        """
        with self._lock:
            now = time.time()
            # 过滤掉已过期的通知
            self._queue = deque(n for n in self._queue if n["expire"] > now)
            return list(self._queue)


# ============================================================
#          DBWriter —— 异步 SQLite 串行写入器
#          （替换原版 AsyncWriter JSON 写入器）
# ============================================================

class DBWriter:
    """
    异步数据库写入器，解决 SQLite 多线程写入竞争问题。

    设计原理：
        SQLite 默认不支持多线程同时写入同一连接，并发写入会抛出
        "database is locked" 异常。

        解决方案：所有写操作通过队列投递到同一个专属写线程（串行执行），
        该线程持有唯一的写连接，零竞争。
        读操作（在主线程）使用独立的读连接，WAL 模式下读写可并发。

    WAL 模式说明：
        Write-Ahead Logging 模式下，写操作先写 WAL 文件，不立即锁定主数据库，
        读操作可以继续访问主数据库，实现读写并发，显著减少锁等待。

    支持的操作类型：
        execute()    → 单条 SQL 写入（非阻塞投递）
        executemany()→ 批量 SQL 写入（非阻塞投递）
        flush()      → 阻塞等待队列清空（程序退出前调用）
        stop()       → 通知写线程优雅退出

    特殊队列消息（内部使用）：
        ("__many__", sql, params_list) → 批量执行
        ("__flush__", event)           → 同步等待落盘
        _STOP 哨兵对象                 → 通知写线程退出
    """

    # 哨兵对象：通知写线程退出主循环
    _STOP = object()

    def __init__(self, db_path: str):
        """
        初始化写入队列并启动后台写线程。

        Args:
            db_path (str): SQLite 数据库文件路径。
        """
        self._db_path = db_path
        self._q: queue.Queue = queue.Queue()  # 无界队列，不丢失写入请求
        self._thread = threading.Thread(
            target=self._worker, daemon=True, name="DBWriter")
        self._thread.start()
        logger.debug("DBWriter 启动，数据库: %s", db_path)

    def execute(self, sql: str, params: tuple = ()):
        """
        投递单条 SQL 写操作（非阻塞，立即返回）。

        Args:
            sql    (str)  : 要执行的 SQL 语句。
            params (tuple): SQL 参数，默认空元组。
        """
        self._q.put((sql, params))

    def executemany(self, sql: str, params_list: list):
        """
        投递批量 SQL 写操作（非阻塞，立即返回）。

        Args:
            sql         (str) : 要批量执行的 SQL 语句。
            params_list (list): 参数列表，每项对应一次执行。
        """
        self._q.put(("__many__", sql, params_list))

    def flush(self, timeout: float = 5.0):
        """
        阻塞等待写队列完全清空（程序退出前调用，确保数据落盘）。

        原理：向队列投递一个带 Event 的 flush 消息，
        写线程处理到该消息时执行 commit 并 set Event，
        调用方的 event.wait() 收到信号后返回。

        Args:
            timeout (float): 最长等待时间（秒），默认 5.0。
        """
        done = threading.Event()
        self._q.put(("__flush__", done))
        done.wait(timeout)

    def stop(self):
        """向写线程投递停止信号，线程处理完当前任务后退出。"""
        self._q.put(self._STOP)

    def _worker(self):
        """
        后台写线程主循环：持有唯一写连接，串行处理所有写操作。

        连接配置：
            - WAL 模式：写操作不阻塞读操作，提升并发性能。
            - synchronous=NORMAL：在 WAL 模式下平衡性能与安全性。

        异常处理：
            execute/executemany 失败时记录 ERROR 日志，
            不向上抛出异常，保证写线程持续运行。
        """
        conn = sqlite3.connect(self._db_path, check_same_thread=False)
        # 启用 WAL 模式：写不阻塞读
        conn.execute("PRAGMA journal_mode=WAL")
        # NORMAL 级别：WAL 模式下足够安全，性能优于 FULL
        conn.execute("PRAGMA synchronous=NORMAL")
        conn.commit()

        while True:
            item = self._q.get()

            # 收到停止哨兵，退出循环
            if item is self._STOP:
                break

            # flush 消息：提交事务并通知调用方
            if item[0] == "__flush__":
                conn.commit()
                item[1].set()   # 唤醒 flush() 的 event.wait()
                continue

            # 批量执行消息
            if item[0] == "__many__":
                _, sql, params_list = item
                try:
                    conn.executemany(sql, params_list)
                    conn.commit()
                except Exception as e:
                    logger.error("DBWriter executemany 失败: %s | SQL: %s", e, sql)
                continue

            # 普通单条执行
            sql, params = item
            try:
                conn.execute(sql, params)
                conn.commit()
            except Exception as e:
                logger.error("DBWriter execute 失败: %s | SQL: %s | params: %s",
                             e, sql, params)

        conn.close()
        logger.debug("DBWriter 已停止")


# ============================================================
#          AttendanceDB —— 数据库访问层（建表 & 读操作）
# ============================================================

class AttendanceDB:
    """
    数据库访问层，负责建表（DDL）和所有只读查询（SELECT）。

    职责划分：
        - 本类只负责建表和读操作，写操作全部经由 DBWriter 异步投递。
        - 读连接（self._rconn）在主线程使用，WAL 模式下与写线程无冲突。

    表结构：
        attendance_status（当前考勤状态，每人一行，按日期区分）：
            name      TEXT PRIMARY KEY   人员姓名
            status    TEXT               已签到 / 已签退
            last_time REAL               最后打卡的 Unix 时间戳（float）
            date      TEXT               YYYY-MM-DD，用于当日数据过滤

        attendance_log（打卡流水，每次打卡一行，永久保留）：
            id        INTEGER PRIMARY KEY AUTOINCREMENT  自增主键
            name      TEXT NOT NULL                      人员姓名
            type      TEXT NOT NULL                      签到 / 签退
            time      TEXT NOT NULL                      YYYY-MM-DD HH:MM:SS
            date      TEXT NOT NULL                      YYYY-MM-DD，冗余存储方便按日查询

    索引说明：
        idx_log_name_date：按（人名 + 日期）查询单人单日记录，O(log n)。
        idx_log_date     ：按日期查询当日所有记录（导出用），O(log n)。
    """

    def __init__(self, db_path: str, writer: DBWriter):
        """
        初始化数据库访问层，建立读连接并执行建表 DDL。

        Args:
            db_path (str)      : SQLite 数据库文件路径。
            writer  (DBWriter) : 异步写入器，供写操作方法使用。
        """
        self._db_path = db_path
        self._writer  = writer
        # 读连接：主线程专用，WAL 模式下与写线程并发无冲突
        self._rconn = sqlite3.connect(db_path, check_same_thread=False)
        self._rconn.execute("PRAGMA journal_mode=WAL")
        # Row 工厂：查询结果支持按列名访问（如 row["name"]）
        self._rconn.row_factory = sqlite3.Row
        self._create_tables()
        logger.info("数据库已就绪: %s", db_path)

    def _create_tables(self):
        """
        创建数据表和索引（IF NOT EXISTS，幂等操作，重复调用无副作用）。

        使用 executescript 批量执行多条 DDL，减少往返次数。
        """
        self._rconn.executescript("""
            CREATE TABLE IF NOT EXISTS attendance_status (
                name      TEXT PRIMARY KEY,
                status    TEXT    NOT NULL DEFAULT '已签到',
                last_time REAL    NOT NULL DEFAULT 0,
                date      TEXT    NOT NULL DEFAULT ''
            );

            CREATE TABLE IF NOT EXISTS attendance_log (
                id   INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT    NOT NULL,
                type TEXT    NOT NULL,
                time TEXT    NOT NULL,
                date TEXT    NOT NULL
            );

            CREATE INDEX IF NOT EXISTS idx_log_name_date
                ON attendance_log(name, date);

            CREATE INDEX IF NOT EXISTS idx_log_date
                ON attendance_log(date);
        """)
        self._rconn.commit()

    # ── 读操作：当日考勤状态 ──

    def get_status(self, name: str, today: str) -> dict | None:
        """
        查询指定人员今日的考勤状态。

        Args:
            name  (str): 人员姓名。
            today (str): 当前日期字符串，格式 "YYYY-MM-DD"。

        Returns:
            dict | None: {"status": str, "last_time": float}，
                         今日尚未打卡则返回 None。
        """
        row = self._rconn.execute(
            "SELECT status, last_time FROM attendance_status "
            "WHERE name=? AND date=?",
            (name, today)
        ).fetchone()
        return dict(row) if row else None

    def get_all_status_today(self, today: str) -> list[dict]:
        """
        查询今天所有已打卡人员的状态列表。

        Args:
            today (str): 当前日期字符串，格式 "YYYY-MM-DD"。

        Returns:
            list[dict]: 每项包含 name、status、last_time 三个字段。
        """
        rows = self._rconn.execute(
            "SELECT name, status, last_time FROM attendance_status WHERE date=?",
            (today,)
        ).fetchall()
        return [dict(r) for r in rows]

    # ── 读操作：打卡日志 ──

    def get_log_today(self, name: str, today: str) -> list[dict]:
        """
        查询指定人员今日的全部打卡记录，按时间升序排列。

        Args:
            name  (str): 人员姓名。
            today (str): 当前日期字符串，格式 "YYYY-MM-DD"。

        Returns:
            list[dict]: 每项包含 type（签到/签退）和 time（时间字符串）。
        """
        rows = self._rconn.execute(
            "SELECT type, time FROM attendance_log "
            "WHERE name=? AND date=? ORDER BY time ASC",
            (name, today)
        ).fetchall()
        return [dict(r) for r in rows]

    def get_log_by_date(self, date: str) -> dict[str, list]:
        """
        查询指定日期所有人员的打卡记录，按时间升序排列。

        用途：Excel 导出时直接从数据库读取，无需依赖内存快照。

        Args:
            date (str): 目标日期字符串，格式 "YYYY-MM-DD"。

        Returns:
            dict[str, list]: 格式 {姓名: [{"type": str, "time": str}, ...]}，
                             与 JSON 版 attendance_log 格式完全兼容。
        """
        rows = self._rconn.execute(
            "SELECT name, type, time FROM attendance_log "
            "WHERE date=? ORDER BY time ASC",
            (date,)
        ).fetchall()
        result: dict[str, list] = {}
        for r in rows:
            result.setdefault(r["name"], []).append(
                {"type": r["type"], "time": r["time"]}
            )
        return result

    def get_recent_records(self, today: str, n: int = 5) -> list[tuple]:
        """
        查询今天最近 n 条打卡流水记录（按时间倒序，直接在 SQL 层排序）。

        Args:
            today (str): 当前日期字符串，格式 "YYYY-MM-DD"。
            n     (int): 返回记录条数，默认 5。

        Returns:
            list[tuple]: 每项格式为 (姓名, 类型, 时间字符串)。
        """
        rows = self._rconn.execute(
            "SELECT name, type, time FROM attendance_log "
            "WHERE date=? ORDER BY time DESC LIMIT ?",
            (today, n)
        ).fetchall()
        return [(r["name"], r["type"], r["time"]) for r in rows]

    # ── 写操作：通过 DBWriter 异步投递（不直接执行）──

    def upsert_status(self, name: str, status: str,
                      last_time: float, today: str):
        """
        异步更新或插入人员考勤状态（INSERT OR REPLACE 语义，幂等）。

        首次打卡：插入新行。
        后续打卡：ON CONFLICT 触发更新，覆盖 status/last_time/date 三个字段。

        Args:
            name      (str)  : 人员姓名。
            status    (str)  : 新的考勤状态（"已签到" 或 "已签退"）。
            last_time (float): 最后打卡的 Unix 时间戳。
            today     (str)  : 当前日期字符串，格式 "YYYY-MM-DD"。
        """
        self._writer.execute(
            """
            INSERT INTO attendance_status(name, status, last_time, date)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(name) DO UPDATE SET
                status    = excluded.status,
                last_time = excluded.last_time,
                date      = excluded.date
            """,
            (name, status, last_time, today)
        )

    def insert_log(self, name: str, action: str,
                   time_str: str, today: str):
        """
        异步插入一条打卡流水记录。

        Args:
            name     (str): 人员姓名。
            action   (str): 打卡动作（"签到" 或 "签退"）。
            time_str (str): 打卡时间字符串，格式 "YYYY-MM-DD HH:MM:SS"。
            today    (str): 当前日期字符串，格式 "YYYY-MM-DD"。
        """
        self._writer.execute(
            "INSERT INTO attendance_log(name, type, time, date) VALUES(?,?,?,?)",
            (name, action, time_str, today)
        )


# ============================================================
#              人脸识别层 —— FaceRecognizer
# ============================================================

class FaceRecognizer:
    """
    职责：加载已知人脸库，对输入 embedding 做相似度匹配，返回身份识别结果。

    与存储层、UI 渲染层完全解耦，只负责"这张脸是谁"这一问题。
    本版与 2.0 版完全一致，无任何修改。
    """

    def __init__(self, known_faces_dir: str = KNOWN_FACES_DIR):
        """
        初始化 InsightFace 模型并加载已知人脸库。

        Args:
            known_faces_dir (str): 已知人员照片目录路径。

        Raises:
            Exception: InsightFace 模型初始化失败时向上抛出，终止程序。
        """
        self._dir     = known_faces_dir
        # 人脸特征数据库：{姓名: 归一化特征向量}
        self.face_db: dict[str, np.ndarray] = {}
        logger.info("初始化 InsightFace ...")
        try:
            self._app = FaceAnalysis(
                name='buffalo_s', providers=['CPUExecutionProvider'])
            self._app.prepare(ctx_id=0, det_size=(320, 320))
        except Exception as e:
            logger.critical("InsightFace 初始化失败: %s", e, exc_info=True)
            raise  # 向上抛出，终止程序启动
        self._build_face_db()
        logger.info("注册人数: %d", len(self.face_db))

    def _build_face_db(self):
        """
        扫描照片目录，逐一提取人脸特征向量，构建内存人脸数据库。

        每张照片单独 try/except，失败只跳过该文件，不影响其他照片加载。
        最终汇总输出成功/跳过数量。
        """
        if not os.path.exists(self._dir):
            os.makedirs(self._dir)
            logger.warning("known_faces 目录不存在，已自动创建: %s", self._dir)
            return
        loaded, failed = 0, 0
        for file in os.listdir(self._dir):
            if not file.lower().endswith(('.jpg', '.png')):
                continue
            name = os.path.splitext(file)[0]
            path = os.path.join(self._dir, file)
            try:
                img = cv2.imread(path)
                if img is None:
                    raise ValueError("cv2.imread 返回 None")
                faces = self._app.get(img)
                if not faces:
                    raise ValueError("图片中未检测到人脸")
                self.face_db[name] = faces[0].normed_embedding
                logger.info("  已注册: %s", name)
                loaded += 1
            except Exception as e:
                logger.warning("  跳过 %s: %s", file, e)
                failed += 1
        logger.info("人脸库加载完成：成功 %d，跳过 %d", loaded, failed)

    def get_faces(self, frame: np.ndarray) -> list:
        """
        对输入帧执行人脸检测与身份识别，返回所有人脸的位置和姓名。

        检测异常时返回空列表，不向上传播，保证主循环不崩溃。

        Args:
            frame (np.ndarray): 输入的 BGR 摄像头帧。

        Returns:
            list: 检测结果列表，每项格式为 (x1, y1, x2, y2, name_or_None)。
        """
        try:
            faces = self._app.get(frame)
        except Exception as e:
            logger.error("人脸检测异常: %s", e, exc_info=True)
            return []
        # 解包 bbox 并附加识别结果，使用解包星号语法简化代码
        return [(*(face.bbox.astype(int)),
                 self._recognize(face.normed_embedding))
                for face in faces]

    def _recognize(self, emb: np.ndarray) -> str | None:
        """
        在人脸数据库中查找与给定特征向量最相似的人员（余弦相似度）。

        face_db 为空时直接返回 None，避免无意义遍历。

        Args:
            emb (np.ndarray): 已归一化的人脸特征向量。

        Returns:
            str | None: 识别到的人员姓名，或 None（未知人员）。
        """
        if not self.face_db:
            return None
        best_name, best_sim = None, -1.0
        for name, db_emb in self.face_db.items():
            sim = float(np.dot(emb, db_emb))  # 归一化向量点积 = 余弦相似度
            if sim > best_sim:
                best_sim, best_name = sim, name
        return best_name if best_sim >= SIMILARITY_THRESHOLD else None


# ============================================================
#     考勤管理层（SQLite 版）—— AttendanceManager
# ============================================================

class AttendanceManager:
    """
    考勤状态与日志的管理层，存储后端从 JSON 文件升级为 SQLite 数据库。

    对外接口（方法签名）与 2.0 版本完全一致，上层代码无需任何修改。

    内存缓存策略（避免每帧都查库）：
        _status_cache：{name: {"status": str, "last_time": float}}
            当日所有已打卡人员的状态，写操作先更新缓存再异步写库。

        _log_cache：{name: [{"type": str, "time": str}, ...]}
            当日所有打卡流水，写操作先追加缓存再异步写库。

        读操作直接读内存缓存，零数据库访问延迟。
        重启恢复：__init__ 时从数据库加载今日数据到内存缓存。

    并发安全：
        所有对 _status_cache / _log_cache 的读写操作通过 RLock 保护。
        写操作（更新缓存 + 投递写库任务）在同一把锁内原子完成，
        消除主线程打卡与导出线程读取之间的竞态条件。
    """

    def __init__(self, db: AttendanceDB):
        """
        初始化考勤管理器，从数据库恢复今日数据到内存缓存。

        Args:
            db (AttendanceDB): 数据库访问层实例。
        """
        self._db    = db
        # RLock：可重入锁，允许同一线程多次获取，防止内部调用死锁
        self._lock  = threading.RLock()
        self._today = datetime.now().strftime("%Y-%m-%d")

        # 内存缓存（当日数据，启动时从数据库恢复）
        self._status_cache: dict[str, dict] = {}
        self._log_cache:    dict[str, list] = {}

        self._load_today()

    def _load_today(self):
        """
        从数据库加载今日考勤数据到内存缓存（启动时调用一次）。

        状态恢复：读取 attendance_status 表中今日的所有行。
        日志恢复：通过 get_log_by_date 读取今日所有打卡流水，按人分组。
        """
        today = self._today
        # 恢复考勤状态缓存
        for row in self._db.get_all_status_today(today):
            self._status_cache[row["name"]] = {
                "status":    row["status"],
                "last_time": row["last_time"]
            }
        # 恢复打卡日志缓存（get_log_by_date 已按人分组）
        log_map = self._db.get_log_by_date(today)
        self._log_cache = log_map
        logger.info("从数据库恢复今日数据：%d 人状态，%d 人日志",
                    len(self._status_cache), len(self._log_cache))

    # ── 打卡核心逻辑 ──

    def do_attendance(self, name: str,
                      notif_queue: NotificationQueue,
                      counter: MultiPersonCounter) -> bool:
        """
        为指定人员执行打卡操作（签到或签退）。

        状态机逻辑：
            - 不在缓存中（今日首次打卡）→ 签到。
            - 已签到状态                 → 签退。
            - 已签退状态                 → 签到（新一轮）。
            - 冷却期内                   → 拒绝，重置帧计数，返回 False。

        并发安全：
            "读缓存→判断冷却→更新缓存→追加日志" 四步在同一把 RLock 内原子完成，
            消除竞态条件，保证不会出现重复打卡或状态错乱。

        性能优化：
            数据库写操作（upsert_status/insert_log）在锁外通过 DBWriter 异步执行，
            不阻塞主线程，锁持有时间最小化。

        Args:
            name        (str)                : 需要打卡的人员姓名。
            notif_queue (NotificationQueue)  : 打卡成功后推送通知的队列。
            counter     (MultiPersonCounter) : 帧计数器，冷却期内需重置。

        Returns:
            bool: True 表示实际发生了打卡；False 表示在冷却期内被跳过。
        """
        now     = time.time()
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        today   = self._today

        with self._lock:
            cached = self._status_cache.get(name)

            if cached is None:
                # 今日首次打卡：直接签到
                action = "签到"
                self._status_cache[name] = {
                    "status":    "已签到",
                    "last_time": now
                }
            else:
                # 冷却期保护：防止同一人被连续快速重复打卡
                if now - cached["last_time"] < COOLDOWN_SECONDS:
                    counter.reset_person(name)  # 重置帧计数，冷却期内不再触发
                    return False
                # 状态切换：已签到→签退，已签退→签到
                action = "签退" if cached["status"] == "已签到" else "签到"
                cached["status"]    = "已签退" if action == "签退" else "已签到"
                cached["last_time"] = now

            # 在锁内追加日志缓存，保证状态与日志的原子一致性
            self._log_cache.setdefault(name, []).append(
                {"type": action, "time": now_str}
            )

        # 锁外异步写库（不阻塞主线程）
        status = self._status_cache[name]["status"]
        self._db.upsert_status(name, status, now, today)
        self._db.insert_log(name, action, now_str, today)

        # 获取工时并推送通知（锁外操作，无需锁保护）
        dur_str, _ = self.get_work_info(name)
        notif_queue.push(name, action, dur_str)
        logger.info("[打卡] %s → %s  %s", name, action, now_str)
        return True

    # ── 查询接口（读内存缓存，零数据库延迟）──

    def get_stats(self) -> tuple:
        """
        获取当日已签到/已签退人数统计（从内存缓存读取）。

        Returns:
            tuple: (signed_in, signed_out)
                - signed_in  (int): 当前处于"已签到"（在岗）状态的人数。
                - signed_out (int): 已完成签退的人数。
        """
        with self._lock:
            si = sum(1 for v in self._status_cache.values()
                     if v["status"] == "已签到")
            so = sum(1 for v in self._status_cache.values()
                     if v["status"] == "已签退")
        return si, so

    def get_work_info(self, name: str) -> tuple:
        """
        获取指定人员当日累计工时和工作段数（从内存缓存读取）。

        锁内复制列表快照，锁外执行计算，减少锁持有时间。

        Args:
            name (str): 人员姓名。

        Returns:
            tuple: (dur_str, segs)
                - dur_str (str): 格式化后的累计工时字符串。
                - segs    (int): 今日完成签到的次数（工作段数）。
        """
        with self._lock:
            recs = list(self._log_cache.get(name, []))  # 锁内复制快照
        _, total, _ = calc_work_sessions(recs)           # 锁外计算
        segs = sum(1 for r in recs if r["type"] == "签到")
        return fmt_dur(total), segs

    def get_status(self, name: str) -> str | None:
        """
        获取指定人员的当前考勤状态（从内存缓存读取）。

        Args:
            name (str): 人员姓名。

        Returns:
            str | None: "已签到" 或 "已签退"，今日未打卡则返回 None。
        """
        with self._lock:
            info = self._status_cache.get(name)
        return info["status"] if info else None

    def get_recent_records(self, n: int = 5) -> list[tuple]:
        """
        获取今日最近 n 条打卡流水记录（从内存缓存汇总并排序）。

        注意：内存缓存中的记录已按追加顺序排列，
        此处重新排序保证多人记录交叉时的时间顺序正确性。

        Args:
            n (int): 返回记录条数，默认 5。

        Returns:
            list[tuple]: 每项格式为 (姓名, 类型, 时间字符串)，按时间倒序。
        """
        with self._lock:
            # 汇总所有人的打卡记录，展开为扁平列表
            all_recs = [
                (nm, e["type"], e["time"])
                for nm, logs in self._log_cache.items()
                for e in logs
            ]
        all_recs.sort(key=lambda x: x[2], reverse=True)  # 按时间倒序
        return all_recs[:n]

    def get_log_for_export(self, date: str | None = None) -> dict:
        """
        导出用：直接从数据库读取指定日期的打卡日志（保证数据最新）。

        与 2.0 版本的 snapshot_log() 不同：
            2.0版：从内存快照导出（可能受内存缓存与数据库不一致影响）。
            3.0版：直接查数据库（数据源唯一，保证最终一致性）。

        Args:
            date (str | None): 目标日期字符串（"YYYY-MM-DD"），
                               None 时默认导出今天的数据。

        Returns:
            dict: 格式 {姓名: [{"type": str, "time": str}, ...]}，
                  与 calc_work_sessions 的输入格式完全兼容。
        """
        target = date or self._today
        return self._db.get_log_by_date(target)


# ============================================================
#             UI 渲染层 —— UIRenderer
#             （与 2.0 版本完全一致，无任何修改）
# ============================================================

class UIRenderer:
    """
    职责：封装所有 OpenCV / PIL 绘制逻辑，不持有考勤业务数据。

    面板分层刷新策略（同 2.0 版本）：
        层1（status 缓存）：打卡统计、最近记录、在场人员，每 0.5s 最多重建一次。
        层2（时钟/FPS）  ：每帧无条件覆盖渲染，彻底消除秒针跳变时的闪烁。
    """

    def __init__(self):
        """初始化面板缓存相关状态变量。"""
        self._panel_status_cache = None   # status 层缓存的图像
        self._panel_status_ts    = 0.0    # status 层最后一次重建的时间戳
        self._last_clock_str     = ""     # 上一帧绘制的时钟字符串（调试用）
        self._last_fps           = -1     # 上一帧绘制的 FPS 值（调试用）

    def draw_faces(self, area, last_faces, sx, sy, cx, cy, manager):
        """
        在视频显示区绘制所有检测到的人脸框、姓名标签、工时及状态信息。

        坐标变换：原始帧坐标 → 视频显示区坐标
            视频区坐标 = 原始坐标 × 缩放比 - 裁剪偏移

        绘制内容（每张人脸）：
            ① 矩形人脸框（已知人员绿色，未知人员黄色）。
            ② 四角 L 形装饰线。
            ③ 姓名标签（框顶部上方，半透明背景）。
            ④ 工时徽章（框底部下方，仅已知人员）。
            ⑤ 签到状态标签（工时徽章下方，仅已打卡人员）。

        Args:
            area       : 视频显示区图像（BGR，原地修改）。
            last_faces : 上一次检测到的人脸列表。
            sx, sy     : X/Y 方向缩放比。
            cx, cy     : X/Y 方向裁剪偏移（像素）。
            manager    : 考勤管理器（用于查询工时和状态）。

        Returns:
            np.ndarray: 绘制人脸信息后的视频区图像。
        """
        text_items = []  # 收集所有文字，最后批量一次 PIL 转换
        for (x1, y1, x2, y2, name) in last_faces:
            # 坐标系变换：原始帧坐标 → 视频显示区坐标
            vx1 = int(x1 * sx) - cx
            vy1 = int(y1 * sy) - cy
            vx2 = int(x2 * sx) - cx
            vy2 = int(y2 * sy) - cy
            # 跳过完全超出视频区边界的人脸框
            if vx2 <= max(0, vx1) or vy2 <= max(0, vy1):
                continue
            col   = C_GREEN if name else C_UNKNOWN
            label = name or "未知人员"
            # ① 矩形人脸框
            cv2.rectangle(area, (vx1, vy1), (vx2, vy2), col, 2)
            # ② 四角 L 形装饰线（增强科技感）
            cl = 18
            for px, py, dx, dy in [(vx1, vy1, 1, 1), (vx2, vy1, -1, 1),
                                    (vx1, vy2, 1, -1), (vx2, vy2, -1, -1)]:
                cv2.line(area, (px, py), (px + dx * cl, py), col, 3)
                cv2.line(area, (px, py), (px, py + dy * cl), col, 3)
            # ③ 姓名标签（人脸框顶部上方）
            ly = max(vy1 - 40, 2)
            blend_rect(area, (vx1, ly - 2),
                       (vx1 + len(label) * 17 + 12, ly + 32), (20, 20, 20), 0.68)
            text_items.append((label, (vx1 + 4, ly), col, 24, True))
            if name:
                # ④ 工时徽章（人脸框底部下方）
                dur_s, segs = manager.get_work_info(name)
                badge = f"工时:{dur_s}({segs}段)"
                blend_rect(area, (vx1, vy2 + 4), (vx1 + 195, vy2 + 30),
                           (20, 20, 20), 0.68)
                text_items.append((badge, (vx1 + 4, vy2 + 6), C_CYAN, 18, False))
                # ⑤ 签到状态标签（工时徽章下方，仅已打卡人员显示）
                st = manager.get_status(name)
                if st:
                    sc  = C_GREEN if st == "已签到" else C_ORANGE
                    sy_ = vy2 + 36
                    blend_rect(area, (vx1, sy_), (vx1 + 88, sy_ + 28),
                               (20, 20, 20), 0.68)
                    text_items.append((st, (vx1 + 4, sy_ + 2), sc, 18, True))
        # 批量执行一次 BGR→PIL→BGR 转换，渲染所有文字
        if text_items:
            area = draw_texts(area, text_items)
        return area

    def draw_notifications(self, area, notif_queue, vw, vh):
        """
        在视频区底部中央绘制打卡成功通知（支持多条同时显示，从下往上堆叠）。

        淡出效果：剩余时间越少，背景透明度越低（alpha = min(0.82, remain * 0.4)）。

        Args:
            area        : 视频显示区图像（BGR，原地修改）。
            notif_queue : 通知队列，提供当前有效通知列表。
            vw          : 视频区宽度（像素）。
            vh          : 视频区高度（像素）。

        Returns:
            np.ndarray: 绘制通知后的视频区图像。
        """
        notifs = notif_queue.get_active()
        if not notifs:
            return area
        bw, bh, gap = 440, 76, 10
        # 从视频区底部往上排列，预留 20px 底部边距
        start_y    = vh - len(notifs) * (bh + gap) - 20
        text_items = []
        for i, notif in enumerate(notifs):
            col    = C_GREEN if notif["action"] == "签到" else C_ORANGE
            # 根据剩余时间动态计算背景透明度（实现淡出效果）
            remain = notif["expire"] - time.time()
            alpha  = min(0.82, remain * 0.4)
            bx     = (vw - bw) // 2  # 水平居中
            by     = start_y + i * (bh + gap)
            # 半透明背景框
            blend_rect(area, (bx, by), (bx + bw, by + bh), (15, 15, 15), alpha)
            # 彩色边框
            cv2.rectangle(area, (bx, by), (bx + bw, by + bh), col, 2)
            # 左侧彩色色条
            cv2.rectangle(area, (bx, by), (bx + 6, by + bh), col, -1)
            text_items += [
                (f"✔  {notif['name']}  {notif['action']}成功",
                 (bx + 14, by + 6),  col,    24, True),   # 主文字
                (f"今日累计工时：{notif['dur_str']}",
                 (bx + 14, by + 40), C_GRAY, 18, False),  # 副文字
            ]
        if text_items:
            area = draw_texts(area, text_items)
        return area

    def draw_panel(self, panel, pw, ph, fps, manager, face_db_count, last_faces):
        """
        渲染右侧信息面板（双层缓存，分离刷新频率）。

        层1 —— status 缓存（每 PANEL_STATUS_TTL 秒最多重建一次）：
            包含：系统标题、日期、今日统计、最近打卡记录、
                  导出按钮提示、当前在场人员、退出提示。

        层2 —— 时钟/FPS（每帧无条件覆盖）：
            每帧强制擦除并重绘时钟和 FPS 区域，彻底消除闪烁。

        姓名截断：按像素宽度截断（_fit_name），支持中英文混排。
        时间戳对齐：预计算像素宽度，动态右对齐。

        Args:
            panel        : 空白面板图像（BGR，将被填充后返回）。
            pw           : 面板宽度（像素）。
            ph           : 面板高度（像素）。
            fps          : 当前帧率（用于 FPS 显示）。
            manager      : 考勤管理器（提供统计和记录数据）。
            face_db_count: 注册人员总数。
            last_faces   : 当前帧检测到的人脸列表（用于在场人员）。

        Returns:
            np.ndarray: 渲染完成的面板图像。
        """
        now    = time.time()
        date_s = datetime.now().strftime("%Y-%m-%d")
        time_s = datetime.now().strftime("%H:%M:%S")

        # ── 层1：status 缓存（打卡统计 + 最近记录 + 在场人员）──
        need_status = (
            self._panel_status_cache is None                      # 首次渲染
            or now - self._panel_status_ts >= PANEL_STATUS_TTL   # 缓存超期
            or self._panel_status_cache.shape[:2] != (ph, pw)    # 窗口尺寸变化
        )

        if need_status:
            # 重建 status 层：完整绘制面板所有静态内容
            panel[:] = C_BG_PANEL
            cv2.rectangle(panel, (0, 0), (pw, 68), (35, 35, 35), -1)
            cv2.line(panel, (0, 68), (pw, 68), C_ACCENT, 2)

            si, so = manager.get_stats()
            nc     = max(0, face_db_count - si - so)  # 未打卡人数

            # 统计数据块配置：(标签, 数值, 颜色, y坐标)
            stats_cfg = [
                ("已签到",   str(si),            C_GREEN,  204),
                ("已签退",   str(so),            C_CYAN,   248),
                ("未打卡",   str(nc),            C_ORANGE, 292),
                ("注册总数", str(face_db_count), C_WHITE,  336),
            ]
            # 统计数据块：半透明背景框 + 左侧彩色色条
            for _, _, col, y in stats_cfg:
                blend_rect(panel, (10, y), (pw - 10, y + 36), (50, 50, 50), 0.5)
                cv2.rectangle(panel, (10, y), (pw - 10, y + 36), C_BORDER, 1)
                cv2.rectangle(panel, (10, y), (14, y + 36), col, -1)

            # 最近 5 条打卡记录（从内存缓存读取）
            recs_all = manager.get_recent_records(5)
            y_rec = 432
            for rn, rt, rtm in recs_all:
                col = C_GREEN if rt == "签到" else C_ORANGE
                blend_rect(panel, (10, y_rec), (pw - 10, y_rec + 44),
                           (50, 50, 50), 0.4)
                cv2.rectangle(panel, (10, y_rec), (pw - 10, y_rec + 44),
                              C_BORDER, 1)
                cv2.rectangle(panel, (10, y_rec), (14, y_rec + 44), col, -1)
                y_rec += 50

            # 导出按钮提示区
            btn_y = y_rec + 8
            blend_rect(panel, (10, btn_y), (pw - 10, btn_y + 36), (0, 80, 50), 0.55)
            cv2.rectangle(panel, (10, btn_y), (pw - 10, btn_y + 36), C_ACCENT, 1)

            # 当前在场人员（本帧视频区检测到的已知人员）
            in_scene = [n for (_, _, _, _, n) in last_faces if n]
            scene_y  = btn_y + 44
            if in_scene:
                cv2.line(panel, (12, scene_y), (pw - 12, scene_y), C_BORDER, 1)
                scene_y += 8

            # ── 姓名像素宽度截断辅助工具 ──
            # 预计算时间戳文字宽度，用于确定姓名可用像素宽度
            _ts_font   = get_font(15, False)
            _ts_sample = "00:00:00"
            try:
                _ts_w = _ts_font.getlength(_ts_sample)      # Pillow 9.2+
            except AttributeError:
                _ts_w = _ts_font.getsize(_ts_sample)[0]     # 旧版 Pillow 兼容

            _name_font   = get_font(17, False)
            # 姓名可用像素宽度 = 面板宽 - 左边距18 - 时间戳宽 - 间距20
            _name_max_px = pw - 18 - int(_ts_w) - 20

            def _fit_name(text, font, max_px):
                """
                按像素宽度截断文字，超出则尾部加 '..'。
                使用二分查找快速定位最大可显示字符数，支持中英文混排。
                """
                try:    w = font.getlength(text)
                except: w = font.getsize(text)[0]
                if w <= max_px:
                    return text
                lo, hi = 0, len(text)
                while lo < hi - 1:
                    mid = (lo + hi) // 2
                    try:    tw = font.getlength(text[:mid] + "..")
                    except: tw = font.getsize(text[:mid] + "..")[0]
                    if tw <= max_px: lo = mid
                    else:            hi = mid
                return text[:lo] + ".."

            # 收集所有静态文字绘制指令
            text_items = [
                ("人脸考勤",     (18, 10),  C_ACCENT,  26, True),
                ("智能管理系统", (14, 38),  C_GRAY,    17, False),
                (date_s,         (10, 82),  C_GRAY,    18, False),
                # 注意：时钟区域留空，将在层2覆盖渲染
                ("▌ 今日统计",   (10, 170), C_ACCENT2, 20, True),
            ]
            # 统计数据块文字：标签 + 数值
            for label, val, col, y in stats_cfg:
                text_items += [
                    (label, (20,      y + 6), C_GRAY, 18, False),
                    (val,   (pw - 38, y + 4), col,    23, True),
                ]
            cv2.line(panel, (12, 384), (pw - 12, 384), C_BORDER, 1)
            text_items.append(("▌ 最近打卡", (10, 396), C_ACCENT2, 20, True))

            # 最近打卡记录文字（姓名按像素截断，时间戳动态右对齐）
            y_rec2 = 432
            for rn, rt, rtm in recs_all:
                col = C_GREEN if rt == "签到" else C_ORANGE
                dn  = _fit_name(rn, _name_font, _name_max_px)
                ts  = rtm[11:]  # 仅取 HH:MM:SS 部分
                text_items += [
                    (dn, (18,                   y_rec2 + 2),  C_WHITE, 17, False),
                    (rt, (18,                   y_rec2 + 24), col,     15, True),
                    (ts, (pw - 10 - int(_ts_w), y_rec2 + 14), C_GRAY, 15, False),
                ]
                y_rec2 += 50

            # 导出按钮和在场人员文字
            text_items.append(("按 E 导出Excel", (16, btn_y + 9), C_ACCENT, 17, True))
            if in_scene:
                text_items.append(("▌ 当前在场", (10, scene_y), C_ACCENT2, 19, True))
                scene_y += 30
                for sname in in_scene[:6]:  # 最多显示 6 人
                    text_items.append((sname, (14, scene_y), C_GREEN, 17, False))
                    scene_y += 24
            # 退出提示（底部固定位置）
            text_items.append(("按 Q 退出", (12, ph - 30), C_GRAY, 16, False))

            # 批量渲染所有静态文字（一次 PIL 转换）
            panel = draw_texts(panel, text_items)
            # 保存 status 层缓存
            self._panel_status_cache = panel.copy()
            self._panel_status_ts    = now
        else:
            # 缓存有效：直接复用 status 层，避免重复 PIL 渲染
            np.copyto(panel, self._panel_status_cache)

        # ── 层2：时钟 + FPS（每帧无条件覆盖，彻底消除闪烁）──
        # status 层重建时会清空面板，若时钟依赖条件触发，
        # 重建后那一帧时钟会短暂消失产生闪烁，解决方案：每帧强制覆盖
        fps_y = ph - 52
        # 擦除时钟和 FPS 区域（用面板背景色覆盖旧内容）
        cv2.rectangle(panel, (0, fps_y - 2), (pw, fps_y + 24), C_BG_PANEL, -1)
        # 覆盖渲染最新时钟和 FPS
        panel = draw_texts(panel, [
            (time_s,        (6,  106),  C_WHITE, 34, True),   # 时钟大字
            (f"FPS: {fps}", (12, fps_y), C_GRAY, 17, False),  # 帧率小字
        ])
        return panel


# ============================================================
#                      Excel 报表导出
# ============================================================

def export_to_excel(attendance_log: dict, face_db_names: list,
                    export_dir: str = EXPORT_DIR) -> str:
    """
    将考勤数据导出为格式化的多 Sheet Excel 报表。

    3.0 版变化：
        attendance_log 参数由调用方直接从数据库读取（manager.get_log_for_export()），
        不再依赖内存快照，数据源更可靠，无并发风险。

    报表结构（4 个 Sheet）：
        Sheet1 - 考勤汇总  ：每人状态、首末打卡时间、总工时、异常说明。
        Sheet2 - 工时明细  ：按工作段展示每段签到/签退时间及累计工时。
        Sheet3 - 原始记录  ：所有打卡流水，按时间升序排列。
        Sheet4 - 异常汇总  ：当日所有考勤异常的聚合视图。

    Args:
        attendance_log (dict): 考勤日志，格式 {姓名: [{"type": str, "time": str}, ...]}。
        face_db_names  (list): 已注册人员姓名列表（用于补全未打卡人员）。
        export_dir     (str) : 报表输出目录，默认 EXPORT_DIR。

    Returns:
        str: 生成的 Excel 文件完整路径。

    Raises:
        OSError: 文件保存失败时向上抛出。
    """
    os.makedirs(export_dir, exist_ok=True)
    today    = datetime.now().strftime("%Y-%m-%d")
    filename = os.path.join(export_dir, f"考勤报表_{today}.xlsx")
    wb       = openpyxl.Workbook()

    # ── 样式定义 ──
    HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")  # 深蓝，表头背景
    SUB_FILL    = PatternFill("solid", fgColor="2E6096")  # 中蓝，合计行
    GREEN_FILL  = PatternFill("solid", fgColor="E2EFDA")  # 浅绿，签到/正常
    ORANGE_FILL = PatternFill("solid", fgColor="FCE4D6")  # 浅橙，签退/未打卡
    GRAY_FILL   = PatternFill("solid", fgColor="F2F2F2")  # 浅灰，交替行
    RED_FILL    = PatternFill("solid", fgColor="FFDDE1")  # 浅红，异常
    BLUE_FILL   = PatternFill("solid", fgColor="DDEEFF")  # 浅蓝，标题行
    WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")  # 白色，正常行
    YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")  # 浅黄，未签退
    DBLUE_FILL  = PatternFill("solid", fgColor="D9E1F2")  # 浅蓝，小计行

    H_FONT     = Font(name="微软雅黑", bold=True,  color="FFFFFF", size=11)  # 表头（白色粗体）
    BOLD_FONT  = Font(name="微软雅黑", bold=True,  size=10)                  # 数据粗体
    NORM_FONT  = Font(name="微软雅黑",             size=10)                  # 普通数据
    SMALL_FONT = Font(name="微软雅黑",             size=9,  color="666666")  # 副标题小字
    RED_FONT   = Font(name="微软雅黑", bold=True,  size=10, color="C00000")  # 红色粗体（异常）
    GREEN_FONT = Font(name="微软雅黑", bold=True,  size=10, color="375623")  # 绿色粗体（正常）
    TITLE_FONT = Font(name="微软雅黑", bold=True,  size=14, color="1E3A5F")  # 大标题

    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    def thin():
        """返回四周统一的细边框样式对象。"""
        s = Side(style="thin", color="CCCCCC")
        return Border(top=s, bottom=s, left=s, right=s)

    def hdr(cell, text, fill=None, font=None):
        """快捷设置表头单元格样式（居中、边框、填充、字体）。"""
        cell.value = text; cell.fill = fill or HEADER_FILL
        cell.font  = font or H_FONT
        cell.alignment = CENTER; cell.border = thin()

    # 合并注册人员与已打卡人员，按姓名排序，确保未打卡人员也出现在报表中
    all_names = sorted(set(list(face_db_names) + list(attendance_log.keys())))

    # ── Sheet1：考勤汇总 ──
    ws1 = wb.active; ws1.title = "考勤汇总"
    # 第1行：报表大标题（跨10列合并）
    ws1.merge_cells("A1:J1")
    c = ws1["A1"]
    c.value = f"📋  员工考勤汇总报表  —  {today}"
    c.font = TITLE_FONT; c.alignment = CENTER; c.fill = BLUE_FILL
    ws1.row_dimensions[1].height = 36
    # 第2行：导出时间 + 注册人数副标题
    ws1.merge_cells("A2:J2")
    c = ws1["A2"]
    c.value = (f"导出时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
               f"    注册人数：{len(face_db_names)}")
    c.font = SMALL_FONT; c.alignment = CENTER; c.fill = GRAY_FILL
    ws1.row_dimensions[2].height = 20

    # 第3行：列标题，同时设置列宽
    s1h = ["序号","姓名","今日状态","首次签到","最后签退",
           "有效工时","工作段数","签到次数","签退次数","异常说明"]
    s1w = [6,14,10,12,12,12,10,10,10,30]
    for ci,(h,w) in enumerate(zip(s1h,s1w),1):
        hdr(ws1.cell(row=3,column=ci),h)
        ws1.column_dimensions[get_column_letter(ci)].width=w
    ws1.row_dimensions[3].height=28

    # 第4行起：逐人填充数据
    row=4
    for idx,name in enumerate(all_names,1):
        records=attendance_log.get(name,[])
        sessions,total_dur,anomalies=calc_work_sessions(records)
        si_cnt=sum(1 for r in records if r["type"]=="签到")
        so_cnt=sum(1 for r in records if r["type"]=="签退")
        first_in=next((s["sign_in"] for s in sessions if s["sign_in"]),"--")
        last_out=next((s["sign_out"] for s in reversed(sessions) if s["sign_out"]),"--")
        # 今日状态判断
        if not records:          status="未打卡"; sf=ORANGE_FILL
        elif any(s["note"]=="未签退" for s in sessions): status="在岗中"; sf=GREEN_FILL
        else:                    status="已完成"; sf=GRAY_FILL
        anom="；".join(anomalies) if anomalies else "正常"
        rf=GRAY_FILL if idx%2==0 else WHITE_FILL  # 斑马纹
        vals=[idx,name,status,first_in,last_out,
              fmt_dur(total_dur),len(sessions),si_cnt,so_cnt,anom]
        for ci,v in enumerate(vals,1):
            cell=ws1.cell(row=row,column=ci,value=v)
            cell.alignment=CENTER if ci!=10 else LEFT; cell.border=thin()
            # 按列应用不同的填充/字体样式
            if ci==3:
                cell.fill=sf
                cell.font=(RED_FONT if status=="未打卡"
                           else GREEN_FONT if status=="在岗中" else BOLD_FONT)
            elif ci==6:
                cell.fill=GREEN_FILL if total_dur.total_seconds()>0 else ORANGE_FILL
                cell.font=GREEN_FONT if total_dur.total_seconds()>0 else RED_FONT
            elif ci==10 and anom!="正常": cell.fill=RED_FILL; cell.font=RED_FONT
            else:
                cell.fill=rf; cell.font=BOLD_FONT if ci==2 else NORM_FONT
        ws1.row_dimensions[row].height=22; row+=1

    # 最后一行：合计/汇总行
    checked=sum(1 for n in all_names if attendance_log.get(n))
    ws1.merge_cells(f"A{row}:B{row}")
    for ci in range(1,11):
        c=ws1.cell(row=row,column=ci)
        c.fill=SUB_FILL; c.font=H_FONT; c.alignment=CENTER; c.border=thin()
    ws1.cell(row=row,column=1).value="合计 / 平均"
    ws1.cell(row=row,column=3).value=f"已打卡:{checked}/{len(all_names)}"
    ws1.row_dimensions[row].height=24; ws1.freeze_panes="A4"

    # ── Sheet2：工时明细 ──
    ws2=wb.create_sheet("工时明细")
    ws2.merge_cells("A1:H1")
    c=ws2["A1"]; c.value=f"📊  工时明细表  —  {today}"
    c.font=TITLE_FONT; c.alignment=CENTER; c.fill=BLUE_FILL
    ws2.row_dimensions[1].height=34
    s2h=["序号","姓名","工作段","签到时间","签退时间","本段工时","状态备注","累计工时"]
    s2w=[6,14,9,14,14,12,14,12]
    for ci,(h,w) in enumerate(zip(s2h,s2w),1):
        hdr(ws2.cell(row=2,column=ci),h)
        ws2.column_dimensions[get_column_letter(ci)].width=w
    ws2.row_dimensions[2].height=26

    drow=3  # 数据起始行号
    for gidx,name in enumerate(all_names,1):
        records=attendance_log.get(name,[])
        sessions,total_dur,_=calc_work_sessions(records)
        # 该人无打卡记录：填充单行占位
        if not sessions:
            for ci,v in enumerate([gidx,name,"--","--","--","--","未打卡","--"],1):
                c=ws2.cell(row=drow,column=ci,value=v)
                c.font=NORM_FONT; c.alignment=CENTER
                c.fill=ORANGE_FILL; c.border=thin()
            ws2.row_dimensions[drow].height=20; drow+=1; continue
        name_start=drow; running_dur=timedelta()  # 记录合并起始行、累计工时
        for sidx,seg in enumerate(sessions,1):
            # 累计有效工时（负数工时不计入）
            if seg["duration"] and seg["duration"].total_seconds()>0:
                running_dur+=seg["duration"]
            # 根据异常类型选择行背景色
            if seg["note"]=="未签退":                     rf2=YELLOW_FILL
            elif seg["note"] in ("无对应签到","时间异常"): rf2=RED_FILL
            else: rf2=WHITE_FILL if sidx%2 else GRAY_FILL  # 斑马纹
            row_vals=[gidx if sidx==1 else "",name if sidx==1 else "",
                      f"第{sidx}段",seg["sign_in"] or "--",seg["sign_out"] or "--",
                      fmt_dur(seg["duration"]),seg["note"] or "正常",fmt_dur(running_dur)]
            for ci,v in enumerate(row_vals,1):
                c=ws2.cell(row=drow,column=ci,value=v)
                c.alignment=CENTER; c.border=thin(); c.fill=rf2
                c.font=(GREEN_FONT if ci==6 and seg["note"]==""
                        else RED_FONT if ci==7 and seg["note"] not in ("","正常")
                        else NORM_FONT)
            ws2.row_dimensions[drow].height=20; drow+=1
        # 小计行：汇总该人所有工作段
        for ci,v in enumerate(["","","小计","","",fmt_dur(total_dur),
                                f"共{len(sessions)}段",fmt_dur(total_dur)],1):
            c=ws2.cell(row=drow,column=ci,value=v)
            c.font=BOLD_FONT; c.fill=DBLUE_FILL; c.alignment=CENTER; c.border=thin()
        ws2.row_dimensions[drow].height=20; drow+=1
        # 合并序号列和姓名列（跨越该人所有工作段行，视觉上归属同一人）
        if drow-2>=name_start:
            try:
                for col in (1,2):
                    ws2.merge_cells(start_row=name_start,start_column=col,
                                    end_row=drow-2,end_column=col)
                    cell=ws2.cell(row=name_start,column=col)
                    cell.alignment=CENTER; cell.font=BOLD_FONT
            except Exception as e:
                # 单行不需要合并时会抛出异常，静默跳过并记录 DEBUG 日志
                logger.debug("合并单元格跳过: %s",e)
    ws2.freeze_panes="A3"

    # ── Sheet3：原始打卡流水 ──
    ws3=wb.create_sheet("原始记录")
    ws3.merge_cells("A1:E1")
    c=ws3["A1"]; c.value=f"📝  原始打卡流水  —  {today}"
    c.font=TITLE_FONT; c.alignment=CENTER; c.fill=BLUE_FILL
    ws3.row_dimensions[1].height=34
    for ci,(h,w) in enumerate(zip(["序号","姓名","类型","打卡时间","备注"],[6,14,10,22,20]),1):
        hdr(ws3.cell(row=2,column=ci),h)
        ws3.column_dimensions[get_column_letter(ci)].width=w
    ws3.row_dimensions[2].height=26
    # 将所有人的打卡记录展开并按时间升序排列
    all_raw=sorted([(n,e["type"],e["time"])
                    for n,logs in attendance_log.items() for e in logs],
                   key=lambda x:x[2])
    for ridx,(rn,rt,rtm) in enumerate(all_raw,1):
        rf3=GREEN_FILL if rt=="签到" else ORANGE_FILL
        for ci,v in enumerate([ridx,rn,rt,rtm,""],1):
            c=ws3.cell(row=ridx+2,column=ci,value=v)
            c.font=(GREEN_FONT if rt=="签到" and ci==3
                    else RED_FONT if rt=="签退" and ci==3 else NORM_FONT)
            c.fill=rf3; c.alignment=CENTER; c.border=thin()
        ws3.row_dimensions[ridx+2].height=20
    ws3.freeze_panes="A3"

    # ── Sheet4：考勤异常汇总 ──
    ws4=wb.create_sheet("异常汇总")
    ws4.merge_cells("A1:D1")
    c=ws4["A1"]; c.value=f"⚠  考勤异常汇总  —  {today}"
    c.font=TITLE_FONT; c.alignment=CENTER
    c.fill=PatternFill("solid",fgColor="FFD7D7")  # 浅红标题背景
    ws4.row_dimensions[1].height=34
    # 列头（使用深红色表头强调异常性质）
    for ci,(h,w) in enumerate(zip(["序号","姓名","异常类型","详细说明"],[6,14,18,40]),1):
        hdr(ws4.cell(row=2,column=ci),h,fill=PatternFill("solid",fgColor="C00000"))
        ws4.column_dimensions[get_column_letter(ci)].width=w
    ws4.row_dimensions[2].height=26
    arow=3; aidx=1; has_anom=False
    for name in all_names:
        _,_,anomalies=calc_work_sessions(attendance_log.get(name,[]))
        for anom in anomalies:
            has_anom=True
            # 异常类型分类（按关键字匹配）
            atype=("未签退" if "未签退" in anom
                   else "无对应签到" if "无对应签到" in anom else "时间异常")
            for ci,v in enumerate([aidx,name,atype,anom],1):
                c=ws4.cell(row=arow,column=ci,value=v)
                c.font=RED_FONT if ci>=3 else NORM_FONT; c.fill=RED_FILL
                c.alignment=LEFT if ci==4 else CENTER; c.border=thin()
            ws4.row_dimensions[arow].height=20; arow+=1; aidx+=1
    # 无异常时显示绿色提示行
    if not has_anom:
        ws4.merge_cells("A3:D3")
        c=ws4.cell(row=3,column=1,value="✅  今日无考勤异常")
        c.font=GREEN_FONT; c.fill=PatternFill("solid",fgColor="E2EFDA")
        c.alignment=CENTER; ws4.row_dimensions[3].height=28

    # 改进3：保存失败时捕获具体异常，记录日志并向上抛出
    try:
        wb.save(filename)
        logger.info("报表已导出：%s", filename)
    except OSError as e:
        logger.error("报表保存失败: %s", e); raise
    return filename


# ============================================================
#          主系统（组合层）—— AttendanceSystem
# ============================================================

class AttendanceSystem:
    """
    考勤管理系统主类（组合层）。

    职责：
        组合 DBWriter、AttendanceDB、FaceRecognizer、
        AttendanceManager、UIRenderer 五个子系统，
        驱动摄像头采集主循环，协调各模块协作，处理键盘事件。

    3.0 版初始化顺序（依赖链）：
        DBWriter → AttendanceDB → AttendanceManager
        FaceRecognizer（独立）
        UIRenderer（独立）

    退出顺序：
        1. 导出最终报表（_do_export）
        2. flush() 等待数据库写入队列清空
        3. stop() 通知写线程优雅退出
        保证程序退出时所有打卡数据已安全落库。
    """

    def __init__(self):
        """
        按依赖顺序初始化所有子系统。
        任何子系统初始化失败都会向上传播，终止程序启动。
        """
        # 初始化顺序：DBWriter → AttendanceDB → AttendanceManager
        self._db_writer  = DBWriter(DB_FILE)           # 异步串行写入器
        self._db         = AttendanceDB(DB_FILE, self._db_writer)  # 数据库访问层
        self.recognizer  = FaceRecognizer()            # 人脸识别层
        self.manager     = AttendanceManager(self._db) # 考勤管理层
        self.renderer    = UIRenderer()                # UI 渲染层
        self._counter    = MultiPersonCounter(CONFIRM_FRAMES)      # 帧计数确认器
        self._notif_q    = NotificationQueue(MAX_NOTIFICATIONS, duration=3.0)  # 通知队列
        self.frame_count = 0         # 帧计数器（用于控制检测频率）
        self.last_faces: list = []   # 上一次检测到的人脸（非检测帧复用）
        logger.info("系统初始化完成 | 注册人数: %d | E=导出  Q=退出",
                    len(self.recognizer.face_db))

    def run(self):
        """
        系统主循环：采集视频帧 → 检测识别 → 渲染 UI → 响应键盘事件。

        帧处理流程：
            1. 读取摄像头帧并水平翻转（镜像显示）。
            2. 每 DETECT_INTERVAL 帧执行一次人脸检测与识别。
            3. MultiPersonCounter 统计连续帧数，达到阈值触发打卡。
            4. crop_fill 将摄像头帧适配到视频显示区尺寸。
            5. 渲染人脸框、通知、信息面板，合并为最终显示帧。

        键盘事件：
            Q → 退出主循环，finally 中依次：导出报表 → flush → stop。
            E → 后台线程异步导出，不阻塞视频流。

        退出保障（finally 块）：
            无论何种退出方式（Q键/Ctrl+C/异常），都保证：
            ① 释放摄像头和销毁窗口。
            ② 导出最终考勤报表。
            ③ 等待数据库写入队列清空（flush）。
            ④ 通知写线程优雅退出（stop）。
        """
        cap = cv2.VideoCapture(0)
        cap.set(cv2.CAP_PROP_FRAME_WIDTH,  CAMERA_WIDTH)
        cap.set(cv2.CAP_PROP_FRAME_HEIGHT, CAMERA_HEIGHT)
        cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)  # 最小缓冲区，降低延迟
        # 摄像头打开失败时优雅退出
        if not cap.isOpened():
            logger.critical("无法打开摄像头，程序退出")
            return

        # 创建全屏窗口
        cv2.namedWindow(WINDOW_NAME, cv2.WINDOW_NORMAL)
        cv2.setWindowProperty(WINDOW_NAME,
                              cv2.WND_PROP_FULLSCREEN,
                              cv2.WINDOW_FULLSCREEN)

        fps_t = time.time(); fps_cnt = fps = 0
        display_buf = None  # 最终合并帧的预分配缓冲区

        try:
            while True:
                ret, frame = cap.read()
                if not ret:
                    # 读帧失败时跳过该帧，不退出循环
                    logger.warning("摄像头读帧失败，跳过")
                    continue
                frame = cv2.flip(frame, 1)  # 水平翻转，实现镜像效果
                orig_h, orig_w = frame.shape[:2]

                # ── FPS 计算 ──
                fps_cnt += 1
                if time.time() - fps_t >= 1.0:
                    fps = fps_cnt; fps_cnt = 0; fps_t = time.time()

                self.frame_count += 1

                # ── 人脸检测与识别（每 DETECT_INTERVAL 帧执行一次）──
                if self.frame_count % DETECT_INTERVAL == 0:
                    self.last_faces = self.recognizer.get_faces(frame)
                    # 使用解包星号语法提取人名集合（跳过 None）
                    detected_names  = {n for (*_, n) in self.last_faces if n}
                    # 帧计数器更新，返回本帧应触发打卡的人名集合
                    for name in self._counter.update(detected_names):
                        if self.manager.do_attendance(
                                name, self._notif_q, self._counter):
                            # 打卡成功后强制使面板 status 层缓存失效
                            self.renderer._panel_status_ts = 0.0

                # ── 获取当前窗口实际尺寸 ──
                try:
                    wr = cv2.getWindowImageRect(WINDOW_NAME)
                    ww, wh = wr[2], wr[3]
                    if ww <= 0 or wh <= 0:
                        raise ValueError
                except Exception:
                    ww, wh = 1920, 1080  # 获取失败时使用默认全屏分辨率

                # ── 计算布局尺寸 ──
                pw = max(240, int(ww * PANEL_RATIO))  # 面板宽度（至少 240px）
                vw = ww - pw                           # 视频区宽度
                vh = wh                                # 视频区高度（同窗口高度）
                # 缩放参数（用于人脸框坐标变换）
                sc = max(vw / orig_w, vh / orig_h)    # crop_fill 缩放比
                cx = (int(orig_w * sc) - vw) // 2     # X 方向裁剪偏移
                cy = (int(orig_h * sc) - vh) // 2     # Y 方向裁剪偏移

                # ── 渲染各区域 ──
                video = self._crop_fill(frame, vw, vh)           # ① 视频区适配
                video = self.renderer.draw_faces(                 # ② 人脸框与标签
                    video, self.last_faces, sc, sc, cx, cy, self.manager)
                video = self.renderer.draw_notifications(         # ③ 打卡通知
                    video, self._notif_q, vw, vh)
                cv2.line(video, (vw - 1, 0), (vw - 1, vh), C_ACCENT, 2)  # ④ 分隔线

                panel = np.zeros((vh, pw, 3), dtype=np.uint8)
                panel = self.renderer.draw_panel(                 # ⑤ 信息面板
                    panel, pw, vh, fps, self.manager,
                    len(self.recognizer.face_db), self.last_faces)

                # ── 合并视频区与面板区（预分配缓冲区，避免每帧内存分配）──
                if display_buf is None or display_buf.shape != (vh, ww, 3):
                    display_buf = np.empty((vh, ww, 3), dtype=np.uint8)
                display_buf[:, :vw]      = video
                display_buf[:, vw:vw+pw] = panel
                cv2.imshow(WINDOW_NAME, display_buf)

                # ── 键盘事件处理 ──
                key = cv2.waitKey(1) & 0xFF
                if key == ord('q'):
                    break  # 退出主循环
                elif key == ord('e'):
                    # 后台线程异步导出，不阻塞视频流
                    threading.Thread(
                        target=self._do_export, daemon=True).start()

        except KeyboardInterrupt:
            # 捕获 Ctrl+C，确保 finally 正常执行资源释放
            logger.info("检测到 Ctrl+C，正在退出...")
        finally:
            # 无论何种退出方式，都保证资源正确释放和数据安全落库
            cap.release()
            cv2.destroyAllWindows()
            logger.info("退出时自动导出报表...")
            self._do_export()          # 同步导出最终报表
            self._db_writer.flush()    # 等待数据库写入队列完全清空
            self._db_writer.stop()     # 通知写线程优雅退出

    @staticmethod
    def _crop_fill(frame, tw, th):
        """
        将输入帧等比缩放并居中裁剪到目标尺寸（类似 CSS background-size: cover）。

        保持宽高比的前提下，取宽/高缩放比中较大的一个，
        缩放后从中心裁剪出 (tw, th) 大小的区域，填满视频显示区。

        Args:
            frame (np.ndarray): 原始摄像头帧（BGR）。
            tw    (int)        : 目标宽度（像素）。
            th    (int)        : 目标高度（像素）。

        Returns:
            np.ndarray: 裁剪后的 (th, tw, 3) BGR 图像。
        """
        fh, fw = frame.shape[:2]
        s  = max(tw / fw, th / fh)    # 取较大缩放比，保证覆盖目标区域
        nw, nh = int(fw * s), int(fh * s)
        r  = cv2.resize(frame, (nw, nh), interpolation=cv2.INTER_LINEAR)
        ox, oy = (nw - tw) // 2, (nh - th) // 2  # 居中裁剪偏移量
        return r[oy:oy + th, ox:ox + tw]

    def _do_export(self):
        """
        执行 Excel 报表导出。

        3.0 版改进：
            直接调用 manager.get_log_for_export() 从数据库读取数据，
            无需内存快照，数据源唯一，保证导出数据的最终一致性。

        可在主线程（退出时同步调用）或后台线程（按 E 时异步调用）中使用。
        捕获所有异常，防止导出失败影响主系统运行。
        """
        try:
            # 直接从数据库读取最新数据，无需内存快照（3.0版改进）
            log_data = self.manager.get_log_for_export()
            export_to_excel(log_data, list(self.recognizer.face_db.keys()))
        except Exception as ex:
            logger.error("导出失败: %s", ex, exc_info=True)


# ============================================================
#                         程序入口
# ============================================================

if __name__ == "__main__":
    system = AttendanceSystem()
    system.run()
