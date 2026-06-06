# -*- coding: utf-8 -*-
"""
人脸识别智能考勤管理系统
========================
功能：
    - 基于 InsightFace 进行实时人脸检测与识别
    - 多人同框同时签到/签退，支持并发打卡
    - 自动计算工时、检测考勤异常
    - 实时 UI 显示（视频区 + 侧边面板）
    - 异步持久化状态与日志
    - 一键导出多 Sheet Excel 考勤报表

依赖：
    insightface, opencv-python, numpy, Pillow,
    openpyxl, xlsxwriter (可选)

使用：
    1. 将已知人员照片（命名为"姓名.jpg/.png"）放入 known_faces/ 目录
    2. 运行本脚本，摄像头自动开启
    3. 按 E 导出 Excel 报表，按 Q 退出
"""

import cv2
import numpy as np
import os
import json
import time
from datetime import datetime, timedelta
from insightface.app import FaceAnalysis
from PIL import Image, ImageDraw, ImageFont
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import threading
import queue
from collections import deque

# ============================================================
#                         全局配置
# ============================================================

# ---------- 文件路径 ----------
KNOWN_FACES_DIR = "known_faces"        # 已知人员照片目录
STATE_FILE      = "attendance_state.json"  # 当日考勤状态持久化文件
LOG_FILE        = "attendance_log.json"    # 打卡流水日志文件
EXPORT_DIR      = "attendance_reports"    # Excel 报表输出目录

# ---------- 识别与打卡参数 ----------
SIMILARITY_THRESHOLD = 0.35   # 人脸相似度阈值，低于此值视为未知人员
CONFIRM_FRAMES       = 6      # 连续检测到同一人至少 N 帧后才触发打卡，防止误识别
COOLDOWN_SECONDS     = 5      # 同一人两次打卡之间的最短冷却时间（秒），防止重复打卡
DETECT_INTERVAL      = 2      # 每隔 N 帧执行一次人脸检测，降低 CPU 占用

# ---------- 摄像头分辨率 ----------
CAMERA_WIDTH  = 1280
CAMERA_HEIGHT = 720

# ---------- UI 字体路径 ----------
FONT_PATH      = "C:/Windows/Fonts/msyh.ttc"    # 微软雅黑常规体
FONT_PATH_BOLD = "C:/Windows/Fonts/msyhbd.ttc"  # 微软雅黑粗体

# ---------- 窗口与布局 ----------
WINDOW_NAME  = "Face Attendance System"  # OpenCV 窗口标题
PANEL_RATIO  = 0.22   # 右侧信息面板占总窗口宽度的比例（0~1）

# ---------- 通知显示 ----------
MAX_NOTIFICATIONS = 4   # 同一时刻最多并排显示的打卡通知条数

# ---------- 颜色常量（BGR 格式） ----------
C_BG_PANEL = (45,  45,  45 )   # 面板背景色（深灰）
C_ACCENT   = (0,   200, 100)   # 主强调色（绿色）
C_ACCENT2  = (0,   160, 255)   # 副强调色（蓝色）
C_WHITE    = (240, 240, 240)   # 近白色
C_GRAY     = (160, 160, 160)   # 灰色，用于次要文字
C_GREEN    = (0,   220, 80 )   # 绿色，表示签到/在岗
C_CYAN     = (200, 220, 0  )   # 青黄色，用于工时信息
C_ORANGE   = (0,   165, 255)   # 橙色，表示签退/警告
C_RED      = (60,  60,  220)   # 红色（BGR），表示错误/异常
C_BORDER   = (80,  80,  80 )   # 边框线颜色
C_UNKNOWN  = (0,   220, 220)   # 黄色，表示未知人员人脸框


# ============================================================
#                       字体缓存模块
# ============================================================

# 全局字体缓存字典，key=(size, bold)，避免重复加载字体文件
_font_cache: dict = {}


def get_font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont:
    """
    获取 PIL 字体对象（带缓存）。

    优先使用粗体路径（FONT_PATH_BOLD），若加载失败则回退至常规体。
    相同 (size, bold) 组合只加载一次，后续直接从缓存取用。

    Args:
        size (int): 字号（像素）。
        bold (bool): 是否使用粗体，默认 False。

    Returns:
        ImageFont.FreeTypeFont: 对应的 PIL 字体对象。
    """
    key = (size, bold)
    if key not in _font_cache:
        try:
            path = FONT_PATH_BOLD if bold else FONT_PATH
            _font_cache[key] = ImageFont.truetype(path, size)
        except Exception:
            # 粗体路径加载失败时，回退使用常规字体路径
            _font_cache[key] = ImageFont.truetype(FONT_PATH, size)
    return _font_cache[key]


# ============================================================
#                       文字渲染模块
# ============================================================

def draw_texts(img_bgr: np.ndarray, items: list) -> np.ndarray:
    """
    批量在 BGR 图像上渲染中文文字（使用 PIL 绘制，避免 OpenCV 乱码）。

    流程：BGR → RGB → PIL 绘制所有文字 → BGR，
    整个批次只做一次颜色空间转换，性能优于逐条调用。

    Args:
        img_bgr (np.ndarray): 输入的 BGR 图像（会被修改后返回）。
        items (list): 文字绘制参数列表，每项格式为：
            (text, (x, y), color_bgr, size, bold)
            - text      (str)  : 要绘制的文字内容
            - (x, y)   (tuple): 文字左上角坐标（像素）
            - color_bgr (tuple): 文字颜色，BGR 格式
            - size      (int)  : 字号（像素）
            - bold      (bool) : 是否粗体

    Returns:
        np.ndarray: 绘制文字后的 BGR 图像。
    """
    if not items:
        return img_bgr
    # 转换到 RGB 以供 PIL 使用
    img_rgb = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2RGB)
    pil_img = Image.fromarray(img_rgb)
    draw    = ImageDraw.Draw(pil_img)
    for text, pos, color_bgr, size, bold in items:
        font = get_font(size, bold)
        # PIL 使用 RGB 颜色，将 BGR 分量反转
        pil_color = (color_bgr[2], color_bgr[1], color_bgr[0])
        draw.text(pos, text, font=font, fill=pil_color)
    # 转换回 BGR 以供 OpenCV 使用
    return cv2.cvtColor(np.array(pil_img), cv2.COLOR_RGB2BGR)


def draw_text(img_bgr, text, pos, color, size, bold=False):
    """
    在 BGR 图像上绘制单条中文文字（兼容旧调用接口）。

    内部复用 draw_texts，保持接口一致性。

    Args:
        img_bgr (np.ndarray): 输入的 BGR 图像。
        text    (str)        : 要绘制的文字内容。
        pos     (tuple)      : 文字左上角坐标 (x, y)。
        color   (tuple)      : 文字颜色，BGR 格式。
        size    (int)        : 字号（像素）。
        bold    (bool)       : 是否粗体，默认 False。

    Returns:
        np.ndarray: 绘制文字后的 BGR 图像。
    """
    return draw_texts(img_bgr, [(text, pos, color, size, bold)])


def blend_rect(img, pt1, pt2, color, alpha=0.55):
    """
    在图像指定区域绘制半透明填充矩形（Alpha 混合）。

    通过 cv2.addWeighted 将纯色矩形与原图 ROI 混合，
    实现毛玻璃/遮罩效果。坐标会被自动裁剪到图像边界内。

    Args:
        img   (np.ndarray): 目标 BGR 图像（原地修改）。
        pt1   (tuple)     : 矩形左上角坐标 (x1, y1)。
        pt2   (tuple)     : 矩形右下角坐标 (x2, y2)。
        color (tuple)     : 填充颜色，BGR 格式。
        alpha (float)     : 遮罩不透明度（0.0=完全透明，1.0=完全不透明），默认 0.55。

    Returns:
        np.ndarray: 混合后的 BGR 图像（与输入共享内存）。
    """
    x1, y1 = pt1; x2, y2 = pt2
    # 坐标裁剪，防止超出图像边界导致越界访问
    x1 = max(0, x1); y1 = max(0, y1)
    x2 = min(img.shape[1]-1, x2)
    y2 = min(img.shape[0]-1, y2)
    if x2 <= x1 or y2 <= y1:
        return img  # 无效区域，直接返回
    roi = img[y1:y2, x1:x2]
    cv2.addWeighted(np.full_like(roi, color), alpha,
                    roi, 1-alpha, 0, roi)
    img[y1:y2, x1:x2] = roi
    return img


# ============================================================
#                      考勤计算模块
# ============================================================

def calc_work_sessions(records: list):
    """
    根据原始打卡流水计算工作段、总工时及异常情况。

    算法：
        - 使用栈（stack_in）配对签到与签退，FIFO 原则。
        - 多余的签退（无对应签到）标记为异常。
        - 未配对的签到（未签退）也标记为异常。

    Args:
        records (list): 打卡记录列表，每项为 dict，格式：
            {"type": "签到"/"签退", "time": "YYYY-MM-DD HH:MM:SS"}

    Returns:
        tuple:
            - sessions  (list)     : 工作段列表，每项含 sign_in、sign_out、
                                     duration (timedelta)、note 字段。
            - total_dur (timedelta): 所有有效工作段的累计工时。
            - anomalies (list)     : 异常描述字符串列表。
    """
    sessions, anomalies, stack_in = [], [], []
    for rec in records:
        # 提取时间部分：支持完整 datetime 字符串或纯时间字符串
        t_str = rec["time"][11:] if len(rec["time"]) > 8 else rec["time"]
        if rec["type"] == "签到":
            stack_in.append(t_str)
        else:
            if stack_in:
                # 取最早的签到与当前签退配对
                sign_in = stack_in.pop(0)
                dur     = _tdiff(sign_in, t_str)
                note    = "" if dur.total_seconds() >= 0 else "时间异常"
                sessions.append({"sign_in": sign_in, "sign_out": t_str,
                                  "duration": dur, "note": note})
                if dur.total_seconds() < 0:
                    anomalies.append(f"签退早于签到:{sign_in}→{t_str}")
            else:
                # 签退时没有对应签到记录
                sessions.append({"sign_in": None, "sign_out": t_str,
                                  "duration": None, "note": "无对应签到"})
                anomalies.append(f"无对应签到:{t_str}")
    # 处理剩余未配对的签到（即未签退）
    for si in stack_in:
        sessions.append({"sign_in": si, "sign_out": None,
                          "duration": None, "note": "未签退"})
        anomalies.append(f"未签退:{si}")
    # 按签到时间升序排列工作段
    sessions.sort(key=lambda s: s["sign_in"] or "99:99:99")
    # 累计所有有效（正数）工作段的工时
    total = sum(
        (s["duration"] for s in sessions
         if s["duration"] and s["duration"].total_seconds() > 0),
        timedelta()
    )
    return sessions, total, anomalies


def _tdiff(t1: str, t2: str) -> timedelta:
    """
    计算两个时间字符串（HH:MM:SS 格式）的差值。

    Args:
        t1 (str): 起始时间字符串，格式 "%H:%M:%S"。
        t2 (str): 结束时间字符串，格式 "%H:%M:%S"。

    Returns:
        timedelta: t2 - t1 的时间差（可能为负数，表示时间异常）。
    """
    fmt = "%H:%M:%S"
    return datetime.strptime(t2, fmt) - datetime.strptime(t1, fmt)


def fmt_dur(td) -> str:
    """
    将 timedelta 对象格式化为人类可读的工时字符串。

    格式规则：
        - None 或无效值 → "--"
        - 负数（异常）  → "异常"
        - 有小时        → "XhYYm"
        - 仅有分钟      → "YmZZs"
        - 仅有秒        → "ZZs"

    Args:
        td (timedelta | None): 时间差对象。

    Returns:
        str: 格式化后的工时字符串。
    """
    if not td: return "--"
    s = int(td.total_seconds())
    if s < 0: return "异常"
    h, m, sec = s//3600, (s%3600)//60, s%60
    if h > 0:   return f"{h}h{m:02d}m"
    elif m > 0: return f"{m}m{sec:02d}s"
    return f"{sec}s"


# ============================================================
#                     多人帧计数器
# ============================================================

class MultiPersonCounter:
    """
    为每个识别到的人独立维护连续帧计数，满足阈值后触发打卡。

    解决原版 "== CONFIRM_FRAMES" 在跳帧场景下可能永远触发不了的 BUG，
    改用 ">=" 判断，并在触发后立即重置计数，避免同一次驻留内重复打卡。

    线程安全：本类未加锁，由调用方保证单线程调用。
    """

    def __init__(self, confirm_frames: int):
        """
        Args:
            confirm_frames (int): 触发打卡所需的最少连续检测帧数。
        """
        self.confirm_frames = confirm_frames
        # 每个人名对应的当前连续检测帧数
        self._counters: dict[str, int] = {}
        # 已触发打卡且尚未离开画面的人名集合（防止在同一驻留周期内重复触发）
        self._triggered: set = set()

    def update(self, detected_names: set) -> set:
        """
        用本帧检测到的人名集合更新内部计数，返回本帧应触发打卡的人名集合。

        逻辑：
            1. 不在本帧的人：计数清零，同时解除触发锁（允许下次出现时重新计数）。
            2. 本帧检测到的人：计数 +1，达到阈值且未触发过则加入返回集合。
            3. 触发后重置该人计数为 0，避免持续触发。

        Args:
            detected_names (set): 本帧识别成功的人名集合。

        Returns:
            set: 本帧需要执行打卡的人名集合（可能为空集）。
        """
        to_attend = set()

        # 消失的人：清零计数并解除触发锁
        for name in list(self._counters.keys()):
            if name not in detected_names:
                self._counters[name] = 0
                self._triggered.discard(name)

        # 本帧检测到的人：累加计数并判断是否达到触发条件
        for name in detected_names:
            self._counters[name] = self._counters.get(name, 0) + 1
            # 使用 >= 而非 == 避免跳帧导致永远无法触发
            if (self._counters[name] >= self.confirm_frames
                    and name not in self._triggered):
                to_attend.add(name)
                self._triggered.add(name)
                # 触发后重置计数，防止下一次签退也被立刻触发
                self._counters[name] = 0

        return to_attend

    def reset_person(self, name: str):
        """
        手动重置指定人员的触发状态（通常在冷却期拒绝打卡后调用）。

        Args:
            name (str): 需要重置状态的人员姓名。
        """
        self._triggered.discard(name)
        self._counters[name] = 0


# ============================================================
#                     多条通知队列
# ============================================================

class NotificationQueue:
    """
    支持多人同框时同时在视频区显示多条打卡通知。

    特性：
        - 每条通知独立计时，到期自动清除（淡出效果由剩余时间控制）。
        - 同一人的旧通知会被新通知替换，避免重复堆叠。
        - 超出最大数量时，移除最旧的通知（FIFO）。
        - 内部使用 threading.Lock 保证多线程安全。
    """

    def __init__(self, max_count: int = 4, duration: float = 3.0):
        """
        Args:
            max_count (int)  : 同时显示的最大通知条数，默认 4。
            duration  (float): 每条通知的显示持续时间（秒），默认 3.0。
        """
        self.max_count = max_count
        self.duration  = duration
        # 通知存储队列，每项为 dict：
        # {"name": str, "action": str, "dur_str": str, "expire": float}
        self._queue: deque = deque()
        self._lock = threading.Lock()

    def push(self, name: str, action: str, dur_str: str):
        """
        推入一条新的打卡通知。

        若同一人已有通知，先移除旧通知再推入新通知。
        若队列已满，移除最旧的通知腾出空间。

        Args:
            name    (str): 打卡人员姓名。
            action  (str): 打卡动作（"签到" 或 "签退"）。
            dur_str (str): 当前累计工时格式化字符串（用于通知显示）。
        """
        with self._lock:
            # 移除同一人的旧通知，防止重复显示
            self._queue = deque(
                n for n in self._queue if n["name"] != name
            )
            if len(self._queue) >= self.max_count:
                self._queue.popleft()  # 移除最旧的通知
            self._queue.append({
                "name":    name,
                "action":  action,
                "dur_str": dur_str,
                "expire":  time.time() + self.duration  # 到期时间戳
            })

    def get_active(self) -> list:
        """
        获取当前所有未过期的通知列表，同时清理已过期的通知。

        Returns:
            list: 未过期通知的有序列表（按推入顺序，最旧的在前）。
        """
        with self._lock:
            now = time.time()
            # 过滤掉已过期的通知
            self._queue = deque(n for n in self._queue
                                if n["expire"] > now)
            return list(self._queue)


# ============================================================
#                      Excel 报表导出
# ============================================================

def export_to_excel(attendance_log: dict, face_db_names: list,
                    export_dir: str = EXPORT_DIR) -> str:
    """
    将当日考勤数据导出为格式化的多 Sheet Excel 报表。

    报表结构（4 个 Sheet）：
        Sheet1 - 考勤汇总：每人今日状态、首末打卡时间、总工时、异常说明。
        Sheet2 - 工时明细：按工作段展示每段签到/签退时间及累计工时。
        Sheet3 - 原始记录：所有打卡流水，按时间排序。
        Sheet4 - 异常汇总：当日所有考勤异常的聚合视图。

    样式说明：
        - 表头使用深蓝填充 + 白色粗体字。
        - 正常行使用白/灰交替填充（斑马纹），异常行使用红色填充。
        - 状态列根据在岗/完成/未打卡使用绿/灰/橙色区分。
        - 所有单元格添加细边框，关键列设置适当列宽。

    Args:
        attendance_log  (dict): 考勤日志字典，格式 {姓名: [打卡记录, ...]}。
        face_db_names   (list): 已注册人员姓名列表（用于补全未打卡人员）。
        export_dir      (str) : 报表输出目录，默认 EXPORT_DIR。

    Returns:
        str: 生成的 Excel 文件完整路径。
    """
    os.makedirs(export_dir, exist_ok=True)
    today    = datetime.now().strftime("%Y-%m-%d")
    filename = os.path.join(export_dir, f"考勤报表_{today}.xlsx")
    wb       = openpyxl.Workbook()

    # ── 样式定义 ──
    # 单元格填充色
    HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")  # 深蓝，表头背景
    SUB_FILL    = PatternFill("solid", fgColor="2E6096")  # 中蓝，合计行背景
    GREEN_FILL  = PatternFill("solid", fgColor="E2EFDA")  # 浅绿，签到/正常
    ORANGE_FILL = PatternFill("solid", fgColor="FCE4D6")  # 浅橙，签退/未打卡
    GRAY_FILL   = PatternFill("solid", fgColor="F2F2F2")  # 浅灰，交替行
    RED_FILL    = PatternFill("solid", fgColor="FFDDE1")  # 浅红，异常
    BLUE_FILL   = PatternFill("solid", fgColor="DDEEFF")  # 浅蓝，标题行
    WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")  # 白色，正常行
    YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")  # 浅黄，未签退
    DBLUE_FILL  = PatternFill("solid", fgColor="D9E1F2")  # 浅蓝，小计行

    # 字体样式
    H_FONT     = Font(name="微软雅黑", bold=True,  color="FFFFFF", size=11)  # 表头字体（白色粗体）
    BOLD_FONT  = Font(name="微软雅黑", bold=True,  size=10)                  # 数据粗体
    NORM_FONT  = Font(name="微软雅黑",             size=10)                  # 普通数据字体
    SMALL_FONT = Font(name="微软雅黑",             size=9, color="666666")   # 小字（副标题）
    RED_FONT   = Font(name="微软雅黑", bold=True,  size=10, color="C00000")  # 红色粗体（异常/警告）
    GREEN_FONT = Font(name="微软雅黑", bold=True,  size=10, color="375623")  # 绿色粗体（正常/签到）
    TITLE_FONT = Font(name="微软雅黑", bold=True,  size=14, color="1E3A5F")  # 大标题字体

    # 对齐方式
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    def thin():
        """返回四周统一的细边框样式对象。"""
        s = Side(style="thin", color="CCCCCC")
        return Border(top=s, bottom=s, left=s, right=s)

    def hdr(cell, text, fill=None, font=None):
        """
        快捷设置表头单元格样式（填充、字体、居中、边框）。

        Args:
            cell        : openpyxl 单元格对象。
            text  (str) : 单元格显示文字。
            fill        : 背景填充，默认使用 HEADER_FILL。
            font        : 字体样式，默认使用 H_FONT。
        """
        cell.value = text
        cell.fill  = fill or HEADER_FILL
        cell.font  = font or H_FONT
        cell.alignment = CENTER
        cell.border    = thin()

    # 合并已注册人员与已打卡人员，按姓名排序，确保未打卡人员也出现在报表中
    all_names = sorted(set(list(face_db_names) + list(attendance_log.keys())))

    # ── Sheet1：考勤汇总 ──
    ws1 = wb.active
    ws1.title = "考勤汇总"

    # 第1行：报表大标题（跨列合并）
    ws1.merge_cells("A1:J1")
    c = ws1["A1"]
    c.value = f"📋  员工考勤汇总报表  —  {today}"
    c.font = TITLE_FONT; c.alignment = CENTER; c.fill = BLUE_FILL
    ws1.row_dimensions[1].height = 36

    # 第2行：导出时间 + 注册人数副标题
    ws1.merge_cells("A2:J2")
    c = ws1["A2"]
    c.value = (f"导出时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
               f"    注册人数：{len(face_db_names)}")
    c.font = SMALL_FONT; c.alignment = CENTER; c.fill = GRAY_FILL
    ws1.row_dimensions[2].height = 20

    # 第3行：列标题，同时设置列宽
    s1h = ["序号","姓名","今日状态","首次签到","最后签退",
           "有效工时","工作段数","签到次数","签退次数","异常说明"]
    s1w = [6, 14, 10, 12, 12, 12, 10, 10, 10, 30]
    for ci, (h, w) in enumerate(zip(s1h, s1w), 1):
        hdr(ws1.cell(row=3, column=ci), h)
        ws1.column_dimensions[get_column_letter(ci)].width = w
    ws1.row_dimensions[3].height = 28

    # 第4行起：逐人填充数据
    row = 4
    for idx, name in enumerate(all_names, 1):
        records  = attendance_log.get(name, [])
        sessions, total_dur, anomalies = calc_work_sessions(records)
        # 统计签到/签退次数
        si_cnt   = sum(1 for r in records if r["type"] == "签到")
        so_cnt   = sum(1 for r in records if r["type"] == "签退")
        # 首次签到时间 / 最后签退时间
        first_in = next((s["sign_in"]  for s in sessions if s["sign_in"]),  "--")
        last_out = next((s["sign_out"] for s in reversed(sessions)
                         if s["sign_out"]), "--")
        # 今日状态判断
        if not records:
            status = "未打卡";  sf = ORANGE_FILL
        elif any(s["note"] == "未签退" for s in sessions):
            status = "在岗中";  sf = GREEN_FILL
        else:
            status = "已完成";  sf = GRAY_FILL

        anom = "；".join(anomalies) if anomalies else "正常"
        # 斑马纹：偶数行用浅灰，奇数行用白色
        rf = GRAY_FILL if idx % 2 == 0 else WHITE_FILL
        vals = [idx, name, status, first_in, last_out,
                fmt_dur(total_dur), len(sessions), si_cnt, so_cnt, anom]
        for ci, v in enumerate(vals, 1):
            cell = ws1.cell(row=row, column=ci, value=v)
            cell.alignment = CENTER if ci != 10 else LEFT
            cell.border    = thin()
            # 按列应用不同的填充/字体样式
            if ci == 3:
                # 状态列：颜色区分
                cell.fill = sf
                cell.font = (RED_FONT   if status == "未打卡"
                             else GREEN_FONT if status == "在岗中"
                             else BOLD_FONT)
            elif ci == 6:
                # 有效工时列：有工时用绿色，否则用橙/红
                cell.fill = GREEN_FILL if total_dur.total_seconds() > 0 else ORANGE_FILL
                cell.font = GREEN_FONT if total_dur.total_seconds() > 0 else RED_FONT
            elif ci == 10 and anom != "正常":
                # 异常说明列：有异常用红色高亮
                cell.fill = RED_FILL; cell.font = RED_FONT
            else:
                cell.fill = rf
                cell.font = BOLD_FONT if ci == 2 else NORM_FONT
        ws1.row_dimensions[row].height = 22
        row += 1

    # 最后一行：合计/汇总行
    checked = sum(1 for n in all_names if attendance_log.get(n))
    ws1.merge_cells(f"A{row}:B{row}")
    for ci in range(1, 11):
        c = ws1.cell(row=row, column=ci)
        c.fill = SUB_FILL; c.font = H_FONT
        c.alignment = CENTER; c.border = thin()
    ws1.cell(row=row, column=1).value = "合计 / 平均"
    ws1.cell(row=row, column=3).value = f"已打卡:{checked}/{len(all_names)}"
    ws1.row_dimensions[row].height = 24
    # 冻结前三行（标题+副标题+列头）
    ws1.freeze_panes = "A4"

    # ── Sheet2：工时明细 ──
    ws2 = wb.create_sheet("工时明细")
    ws2.merge_cells("A1:H1")
    c = ws2["A1"]
    c.value = f"📊  工时明细表  —  {today}"
    c.font = TITLE_FONT; c.alignment = CENTER; c.fill = BLUE_FILL
    ws2.row_dimensions[1].height = 34

    # 列头及列宽设置
    s2h = ["序号","姓名","工作段","签到时间","签退时间","本段工时","状态备注","累计工时"]
    s2w = [6, 14, 9, 14, 14, 12, 14, 12]
    for ci, (h, w) in enumerate(zip(s2h, s2w), 1):
        hdr(ws2.cell(row=2, column=ci), h)
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.row_dimensions[2].height = 26

    drow = 3  # 数据起始行号
    for gidx, name in enumerate(all_names, 1):
        records  = attendance_log.get(name, [])
        sessions, total_dur, _ = calc_work_sessions(records)

        # 该人无打卡记录：填充单行占位
        if not sessions:
            for ci, v in enumerate([gidx, name, "--", "--", "--", "--", "未打卡", "--"], 1):
                c = ws2.cell(row=drow, column=ci, value=v)
                c.font = NORM_FONT; c.alignment = CENTER
                c.fill = ORANGE_FILL; c.border = thin()
            ws2.row_dimensions[drow].height = 20; drow += 1; continue

        name_start  = drow     # 记录姓名列合并的起始行
        running_dur = timedelta()  # 当前累计工时（逐段累加）

        for sidx, seg in enumerate(sessions, 1):
            # 累计有效工时
            if seg["duration"] and seg["duration"].total_seconds() > 0:
                running_dur += seg["duration"]
            # 根据异常类型选择行背景色
            if seg["note"] == "未签退":
                rf2 = YELLOW_FILL
            elif seg["note"] in ("无对应签到", "时间异常"):
                rf2 = RED_FILL
            else:
                rf2 = WHITE_FILL if sidx % 2 else GRAY_FILL  # 斑马纹

            row_vals = [
                gidx if sidx == 1 else "",   # 序号只在第一段显示
                name if sidx == 1 else "",   # 姓名只在第一段显示
                f"第{sidx}段",
                seg["sign_in"]  or "--",
                seg["sign_out"] or "--",
                fmt_dur(seg["duration"]),
                seg["note"] or "正常",
                fmt_dur(running_dur)          # 截至本段的累计工时
            ]
            for ci, v in enumerate(row_vals, 1):
                c = ws2.cell(row=drow, column=ci, value=v)
                c.alignment = CENTER; c.border = thin(); c.fill = rf2
                c.font = (GREEN_FONT if ci == 6 and seg["note"] == ""
                          else RED_FONT if ci == 7 and seg["note"] not in ("", "正常")
                          else NORM_FONT)
            ws2.row_dimensions[drow].height = 20; drow += 1

        # 小计行：汇总该人所有工作段
        for ci, v in enumerate(["", "", "小计", "", "",
                                 fmt_dur(total_dur),
                                 f"共{len(sessions)}段",
                                 fmt_dur(total_dur)], 1):
            c = ws2.cell(row=drow, column=ci, value=v)
            c.font = BOLD_FONT; c.fill = DBLUE_FILL
            c.alignment = CENTER; c.border = thin()
        ws2.row_dimensions[drow].height = 20; drow += 1

        # 合并序号列和姓名列（跨越该人所有工作段行），视觉上归属同一人
        if drow - 2 >= name_start:
            try:
                for col in (1, 2):
                    ws2.merge_cells(start_row=name_start, start_column=col,
                                    end_row=drow-2, end_column=col)
                    cell = ws2.cell(row=name_start, column=col)
                    cell.alignment = CENTER; cell.font = BOLD_FONT
            except Exception:
                pass  # 合并失败时静默跳过（如单行不需要合并）

    ws2.freeze_panes = "A3"

    # ── Sheet3：原始打卡流水 ──
    ws3 = wb.create_sheet("原始记录")
    ws3.merge_cells("A1:E1")
    c = ws3["A1"]
    c.value = f"📝  原始打卡流水  —  {today}"
    c.font = TITLE_FONT; c.alignment = CENTER; c.fill = BLUE_FILL
    ws3.row_dimensions[1].height = 34

    # 列头设置
    for ci, (h, w) in enumerate(zip(["序号","姓名","类型","打卡时间","备注"],
                                     [6, 14, 10, 22, 20]), 1):
        hdr(ws3.cell(row=2, column=ci), h)
        ws3.column_dimensions[get_column_letter(ci)].width = w
    ws3.row_dimensions[2].height = 26

    # 将所有人的打卡记录展开并按时间升序排列
    all_raw = sorted(
        [(n, e["type"], e["time"])
         for n, logs in attendance_log.items() for e in logs],
        key=lambda x: x[2]
    )
    for ridx, (rn, rt, rtm) in enumerate(all_raw, 1):
        rf3 = GREEN_FILL if rt == "签到" else ORANGE_FILL
        for ci, v in enumerate([ridx, rn, rt, rtm, ""], 1):
            c = ws3.cell(row=ridx+2, column=ci, value=v)
            c.font = (GREEN_FONT if rt == "签到"  and ci == 3
                      else RED_FONT if rt == "签退" and ci == 3
                      else NORM_FONT)
            c.fill = rf3; c.alignment = CENTER; c.border = thin()
        ws3.row_dimensions[ridx+2].height = 20
    ws3.freeze_panes = "A3"

    # ── Sheet4：考勤异常汇总 ──
    ws4 = wb.create_sheet("异常汇总")
    ws4.merge_cells("A1:D1")
    c = ws4["A1"]
    c.value = f"⚠  考勤异常汇总  —  {today}"
    c.font = TITLE_FONT; c.alignment = CENTER
    c.fill = PatternFill("solid", fgColor="FFD7D7")  # 浅红标题背景
    ws4.row_dimensions[1].height = 34

    # 列头（使用深红色表头强调异常性质）
    for ci, (h, w) in enumerate(zip(["序号","姓名","异常类型","详细说明"],
                                     [6, 14, 18, 40]), 1):
        hdr(ws4.cell(row=2, column=ci), h,
            fill=PatternFill("solid", fgColor="C00000"))  # 深红表头
        ws4.column_dimensions[get_column_letter(ci)].width = w
    ws4.row_dimensions[2].height = 26

    arow = 3; aidx = 1; has_anom = False
    for name in all_names:
        _, _, anomalies = calc_work_sessions(attendance_log.get(name, []))
        for anom in anomalies:
            has_anom = True
            # 异常类型分类
            atype = ("未签退"    if "未签退"    in anom
                     else "无对应签到" if "无对应签到" in anom
                     else "时间异常")
            for ci, v in enumerate([aidx, name, atype, anom], 1):
                c = ws4.cell(row=arow, column=ci, value=v)
                c.font = RED_FONT if ci >= 3 else NORM_FONT
                c.fill = RED_FILL
                c.alignment = LEFT if ci == 4 else CENTER
                c.border = thin()
            ws4.row_dimensions[arow].height = 20
            arow += 1; aidx += 1

    # 无异常时显示提示行
    if not has_anom:
        ws4.merge_cells("A3:D3")
        c = ws4.cell(row=3, column=1, value="✅  今日无考勤异常")
        c.font = GREEN_FONT
        c.fill = PatternFill("solid", fgColor="E2EFDA")
        c.alignment = CENTER
        ws4.row_dimensions[3].height = 28

    wb.save(filename)
    print(f"✅ 报表已导出：{filename}")
    return filename


# ============================================================
#                      异步文件写入器
# ============================================================

class AsyncWriter:
    """
    基于后台线程 + 队列的异步 JSON 文件写入器。

    打卡触发时，主线程只需将数据放入队列即可立即返回，
    不会因磁盘 I/O 阻塞摄像头采集帧。

    设计说明：
        - 后台线程为守护线程，随主进程退出自动销毁。
        - 队列无限大，不会丢失写入请求。
        - 若同一文件被多次推入，最终以最后一次为准（后写覆盖前写）。
    """

    def __init__(self):
        """初始化写入队列并启动后台工作线程。"""
        self._q = queue.Queue()
        threading.Thread(target=self._worker, daemon=True).start()

    def write(self, path: str, data: dict):
        """
        将写入任务放入队列（非阻塞）。

        Args:
            path (str) : 目标文件路径。
            data (dict): 要序列化写入的 Python 字典。
        """
        self._q.put((path, data))

    def _worker(self):
        """
        后台工作线程主循环：持续从队列取出任务并执行文件写入。

        写入失败时打印错误信息，不抛出异常，保证线程持续运行。
        """
        while True:
            path, data = self._q.get()
            try:
                with open(path, "w", encoding="utf-8") as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
            except Exception as e:
                print(f"[Writer] 写入失败: {e}")


# ============================================================
#                        主考勤系统
# ============================================================

class AttendanceSystem:
    """
    人脸识别智能考勤管理系统主类。

    职责：
        - 初始化 InsightFace 模型、加载人员数据库。
        - 管理摄像头输入与视频帧处理循环。
        - 协调人脸检测、识别、帧计数确认、打卡记录等流程。
        - 渲染实时 UI（视频区人脸框 + 右侧信息面板 + 打卡通知）。
        - 将考勤状态与日志持久化到 JSON 文件。
        - 响应键盘事件（E=导出、Q=退出）。
    """

    def __init__(self):
        """
        系统初始化：加载模型 → 注册人脸库 → 恢复当日状态与日志。
        """
        print("初始化 InsightFace ...")
        # 使用轻量级 buffalo_s 模型，CPU 推理
        self.app = FaceAnalysis(
            name='buffalo_s',
            providers=['CPUExecutionProvider']
        )
        self.app.prepare(ctx_id=0, det_size=(320, 320))

        # 人脸特征数据库：{姓名: 归一化特征向量}
        self.face_db           = {}
        # 当日考勤状态：{姓名: {"status": "已签到"/"已签退", "last_time": timestamp}}
        self.attendance_status = {}
        # 打卡流水日志：{姓名: [{"type": "签到"/"签退", "time": "..."}]}
        self.attendance_log    = {}
        # 帧计数器（用于控制检测频率）
        self.frame_count       = 0
        # 上一次检测到的人脸列表：[(x1,y1,x2,y2,name), ...]
        # 非检测帧复用上次结果，保持绘制连续性
        self.last_faces        = []

        # 多人帧计数确认器（修复原版跳帧 BUG）
        self._counter = MultiPersonCounter(CONFIRM_FRAMES)

        # 多条打卡通知队列（支持多人同框同时显示通知）
        self._notif_queue = NotificationQueue(
            max_count=MAX_NOTIFICATIONS, duration=3.0)

        # 异步 JSON 文件写入器（避免磁盘 I/O 阻塞主线程）
        self._writer = AsyncWriter()

        # 右侧面板缓存：在内容未变化时复用上次渲染结果，减少 PIL 转换开销
        self._panel_cache    = None    # 缓存的面板图像（numpy array）
        self._panel_cache_ts = 0.0     # 缓存生成时间戳
        self._PANEL_TTL      = 0.5     # 面板缓存有效期（秒），超出后强制重绘

        self._build_face_db()
        self._load_state()
        self._load_log()
        print(f"✅ 系统初始化完成，注册人数: {len(self.face_db)}")
        print("E=导出报表  Q=退出")

    # ──────────────────────────────────────────────
    #                  数据加载与持久化
    # ──────────────────────────────────────────────

    def _build_face_db(self):
        """
        扫描 KNOWN_FACES_DIR 目录，提取所有已知人员的人脸特征向量，构建内存数据库。

        命名规则：文件名（不含扩展名）即为人员姓名。
        支持格式：.jpg、.png（大小写不限）。
        若图像中无法检测到人脸，则跳过该文件并打印警告。
        """
        if not os.path.exists(KNOWN_FACES_DIR):
            os.makedirs(KNOWN_FACES_DIR); return
        for file in os.listdir(KNOWN_FACES_DIR):
            if not file.lower().endswith(('.jpg', '.png')): continue
            name = os.path.splitext(file)[0]
            img  = cv2.imread(os.path.join(KNOWN_FACES_DIR, file))
            if img is None: continue
            faces = self.app.get(img)
            if faces:
                self.face_db[name] = faces[0].normed_embedding
                print(f"  已注册: {name}")

    def _load_state(self):
        """
        从 STATE_FILE 恢复当日考勤状态（仅当日数据有效）。

        若文件中的日期与当前日期不符，则忽略（系统自然重置为新一天）。
        """
        today = datetime.now().strftime("%Y-%m-%d")
        if os.path.exists(STATE_FILE):
            with open(STATE_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            if data.get("date") == today:
                self.attendance_status = data["attendance_status"]

    def _save_state(self):
        """
        将当前考勤状态异步写入 STATE_FILE。

        每次打卡成功后调用，通过 AsyncWriter 队列异步落盘。
        """
        today = datetime.now().strftime("%Y-%m-%d")
        self._writer.write(STATE_FILE, {
            "date":               today,
            "attendance_status":  self.attendance_status
        })

    def _load_log(self):
        """
        从 LOG_FILE 恢复当日打卡流水日志（仅当日数据有效）。

        若文件中的日期与当前日期不符，则忽略。
        """
        today = datetime.now().strftime("%Y-%m-%d")
        if os.path.exists(LOG_FILE):
            with open(LOG_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            if data.get("date") == today:
                self.attendance_log = data.get("log", {})

    def _save_log(self):
        """
        将当前打卡流水日志异步写入 LOG_FILE。

        每次打卡成功后调用，通过 AsyncWriter 队列异步落盘。
        """
        today = datetime.now().strftime("%Y-%m-%d")
        self._writer.write(LOG_FILE, {
            "date": today,
            "log":  self.attendance_log
        })

    # ──────────────────────────────────────────────
    #                    人脸识别
    # ──────────────────────────────────────────────

    def _recognize(self, emb) -> str | None:
        """
        在人脸数据库中查找与给定特征向量最相似的人员。

        算法：计算输入特征与所有注册特征的余弦相似度（归一化向量点积），
        取最大值，若超过 SIMILARITY_THRESHOLD 则返回对应姓名，否则返回 None。

        Args:
            emb (np.ndarray): 已归一化的人脸特征向量（来自 InsightFace）。

        Returns:
            str | None: 识别到的人员姓名，或 None（未知人员）。
        """
        best_name, best_sim = None, -1.0
        for name, db_emb in self.face_db.items():
            sim = float(np.dot(emb, db_emb))
            if sim > best_sim:
                best_sim = sim; best_name = name
        return best_name if best_sim >= SIMILARITY_THRESHOLD else None

    # ──────────────────────────────────────────────
    #                    打卡逻辑
    # ──────────────────────────────────────────────

    def _do_attendance(self, name: str):
        """
        为指定人员执行打卡操作（签到或签退）。

        状态机逻辑：
            - 首次打卡（不在状态表中） → 签到。
            - 已签到状态 → 签退。
            - 已签退状态 → 签到（新一轮）。
            - 冷却期内（距上次打卡 < COOLDOWN_SECONDS） → 拒绝并重置计数器。

        打卡成功后：
            1. 更新内存中的状态表和日志。
            2. 推送通知到 NotificationQueue。
            3. 使面板缓存失效（强制重绘）。
            4. 异步写入状态文件和日志文件。

        Args:
            name (str): 需要打卡的人员姓名（已通过识别确认）。
        """
        now     = time.time()
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if name not in self.attendance_status:
            # 该人今日首次打卡：直接签到
            self.attendance_status[name] = {
                "status": "已签到", "last_time": now}
            action = "签到"
        else:
            info = self.attendance_status[name]
            # 冷却期保护：防止同一人被连续快速打卡
            if now - info["last_time"] < COOLDOWN_SECONDS:
                self._counter.reset_person(name)  # 重置帧计数，冷却期内不再触发
                return
            # 状态切换：已签到→签退，已签退→签到
            action              = "签退" if info["status"] == "已签到" else "签到"
            info["status"]    = "已签退" if action == "签退" else "已签到"
            info["last_time"] = now

        # 记录到流水日志
        self.attendance_log.setdefault(name, []).append(
            {"type": action, "time": now_str})

        # 获取最新累计工时，推入通知队列
        dur_str, _ = self._get_work_info(name)
        self._notif_queue.push(name, action, dur_str)

        # 使面板缓存失效，下一帧强制重绘面板
        self._panel_cache_ts = 0.0

        # 异步落盘
        self._save_state()
        self._save_log()
        print(f"[打卡] {name} → {action}  {now_str}")

    # ──────────────────────────────────────────────
    #                    统计工具
    # ──────────────────────────────────────────────

    def _get_stats(self):
        """
        获取当日签到/签退/未打卡人数统计。

        Returns:
            tuple: (total, signed_in, signed_out, not_checked)
                - total       (int): 注册人员总数。
                - signed_in   (int): 当前处于"已签到"（在岗）状态的人数。
                - signed_out  (int): 已完成签退的人数。
                - not_checked (int): 尚未任何打卡记录的人数。
        """
        total = len(self.face_db)
        si    = sum(1 for v in self.attendance_status.values()
                    if v["status"] == "已签到")
        so    = sum(1 for v in self.attendance_status.values()
                    if v["status"] == "已签退")
        nc    = total - len(self.attendance_status)
        return total, si, so, nc

    def _get_work_info(self, name: str):
        """
        获取指定人员当日累计工时和工作段数（用于 UI 显示）。

        Args:
            name (str): 人员姓名。

        Returns:
            tuple: (dur_str, segs)
                - dur_str (str): 格式化后的累计工时字符串，如 "2h30m"。
                - segs    (int): 今日完成签到的次数（工作段数）。
        """
        recs = self.attendance_log.get(name, [])
        _, total, _ = calc_work_sessions(recs)
        segs = sum(1 for r in recs if r["type"] == "签到")
        return fmt_dur(total), segs

    # ──────────────────────────────────────────────
    #                   视频区处理
    # ──────────────────────────────────────────────

    def _crop_fill(self, frame, tw, th):
        """
        将输入帧缩放并居中裁剪到目标尺寸（类似 CSS background-size: cover）。

        保持宽高比的前提下，取宽/高缩放比中较大的一个，
        缩放后从中心裁剪出 (tw, th) 大小的区域，填满视频显示区。

        Args:
            frame (np.ndarray): 原始摄像头帧（BGR）。
            tw    (int)        : 目标宽度（像素）。
            th    (int)        : 目标高度（像素）。

        Returns:
            np.ndarray: 裁剪后的 (th, tw, 3) BGR 图像。
        """
        fh, fw = frame.shape[:2]
        s  = max(tw/fw, th/fh)   # 取较大缩放比，保证覆盖目标区域
        nw, nh = int(fw*s), int(fh*s)
        r  = cv2.resize(frame, (nw, nh), interpolation=cv2.INTER_LINEAR)
        ox, oy = (nw-tw)//2, (nh-th)//2  # 居中裁剪偏移量
        return r[oy:oy+th, ox:ox+tw]

    def _draw_faces(self, area, sx, sy, cx, cy):
        """
        在视频显示区绘制所有检测到的人脸框、标签、工时及状态信息。

        坐标转换说明：
            人脸检测在原始帧坐标系下进行，需映射到经过 crop_fill 处理后的
            视频显示区坐标系。转换公式：
                视频区坐标 = 原始坐标 * 缩放比 - 裁剪偏移

        绘制内容（每张人脸）：
            1. 矩形人脸框（已知人员绿色，未知人员黄色）。
            2. 四角装饰线（L 形角标）。
            3. 姓名标签（半透明背景 + 文字）。
            4. 工时徽章（已知人员）。
            5. 签到状态标签（已知人员且已打卡）。

        Args:
            area (np.ndarray): 视频显示区图像（BGR，原地修改）。
            sx   (float)     : X 方向缩放比（视频区宽度 / 原始帧宽度）。
            sy   (float)     : Y 方向缩放比（视频区高度 / 原始帧高度）。
            cx   (int)       : X 方向裁剪偏移（像素）。
            cy   (int)       : Y 方向裁剪偏移（像素）。

        Returns:
            np.ndarray: 绘制人脸信息后的视频区图像。
        """
        text_items = []  # 收集所有需要渲染的文字，最后批量一次 PIL 转换

        for (x1, y1, x2, y2, name) in self.last_faces:
            # 坐标系变换：原始帧坐标 → 视频显示区坐标
            vx1 = int(x1*sx) - cx; vy1 = int(y1*sy) - cy
            vx2 = int(x2*sx) - cx; vy2 = int(y2*sy) - cy

            # 限制坐标在画面内，防止越界绘制
            H, W = area.shape[:2]
            vx1c = max(0, vx1); vy1c = max(0, vy1)
            vx2c = min(W, vx2); vy2c = min(H, vy2)
            if vx2c <= vx1c or vy2c <= vy1c:
                continue  # 人脸框完全超出视频区，跳过

            # 颜色与标签：已知人员绿色，未知人员黄色
            col   = C_GREEN if name else C_UNKNOWN
            label = name or "未知人员"

            # ① 矩形人脸框
            cv2.rectangle(area, (vx1, vy1), (vx2, vy2), col, 2)

            # ② 四角 L 形装饰线（增强科技感）
            cl = 18  # 角标长度（像素）
            for px, py, dx, dy in [(vx1, vy1,  1,  1), (vx2, vy1, -1,  1),
                                    (vx1, vy2,  1, -1), (vx2, vy2, -1, -1)]:
                cv2.line(area, (px, py), (px+dx*cl, py), col, 3)  # 水平线
                cv2.line(area, (px, py), (px, py+dy*cl), col, 3)  # 垂直线

            # ③ 姓名标签（框顶部上方，半透明背景）
            ly = max(vy1 - 40, 2)
            blend_rect(area, (vx1, ly-2),
                       (vx1 + len(label)*17 + 12, ly+32), (20, 20, 20), 0.68)
            text_items.append((label, (vx1+4, ly), col, 24, True))

            if name:
                # ④ 工时徽章（框底部下方）
                dur_s, segs = self._get_work_info(name)
                badge = f"工时:{dur_s}({segs}段)"
                blend_rect(area, (vx1, vy2+4), (vx1+195, vy2+30),
                           (20, 20, 20), 0.68)
                text_items.append((badge, (vx1+4, vy2+6), C_CYAN, 18, False))

                # ⑤ 签到状态标签（工时徽章下方）
                if name in self.attendance_status:
                    st  = self.attendance_status[name]["status"]
                    sc  = C_GREEN if st == "已签到" else C_ORANGE
                    sy_ = vy2 + 36
                    blend_rect(area, (vx1, sy_), (vx1+88, sy_+28),
                               (20, 20, 20), 0.68)
                    text_items.append((st, (vx1+4, sy_+2), sc, 18, True))

        # 批量执行一次 BGR→PIL→BGR 转换，渲染所有文字
        if text_items:
            area = draw_texts(area, text_items)
        return area

    def _draw_notifications(self, area, vw, vh):
        """
        在视频区底部中央绘制打卡成功通知（支持多条同时显示，从下往上堆叠）。

        通知样式：
            - 半透明深色背景矩形框，左侧带彩色色条（签到绿/签退橙）。
            - 顶行：打卡人姓名 + 动作（粗体彩色）。
            - 底行：今日累计工时（灰色小字）。
            - 淡出效果：通知剩余时间越短，背景透明度越低。

        Args:
            area (np.ndarray): 视频显示区图像（BGR，原地修改）。
            vw   (int)        : 视频区宽度（像素）。
            vh   (int)        : 视频区高度（像素）。

        Returns:
            np.ndarray: 绘制通知后的视频区图像。
        """
        notifs = self._notif_queue.get_active()
        if not notifs:
            return area

        bw, bh = 440, 76   # 单条通知框的宽度和高度
        gap    = 10         # 通知框之间的垂直间距
        # 计算所有通知框堆叠的总高度，从视频区底部往上排列
        total_h = len(notifs) * (bh + gap)
        start_y = vh - total_h - 20

        text_items = []  # 收集所有文字，批量渲染

        for i, notif in enumerate(notifs):
            name    = notif["name"]
            action  = notif["action"]
            dur_str = notif["dur_str"]
            # 根据剩余显示时间动态计算背景透明度（实现淡出效果）
            remain  = notif["expire"] - time.time()
            alpha   = min(0.82, remain * 0.4)  # 最大 0.82，线性淡出

            col = C_GREEN if action == "签到" else C_ORANGE  # 签到绿/签退橙
            bx  = (vw - bw) // 2   # 水平居中
            by  = start_y + i * (bh + gap)

            # 半透明背景框
            blend_rect(area, (bx, by), (bx+bw, by+bh), (15, 15, 15), alpha)
            # 彩色边框
            cv2.rectangle(area, (bx, by), (bx+bw, by+bh), col, 2)
            # 左侧彩色色条
            cv2.rectangle(area, (bx, by), (bx+6, by+bh), col, -1)

            text_items += [
                (f"✔  {name}  {action}成功",
                 (bx+14, by+6),  col,    24, True),   # 主文字
                (f"今日累计工时：{dur_str}",
                 (bx+14, by+40), C_GRAY, 18, False),  # 副文字
            ]

        if text_items:
            area = draw_texts(area, text_items)
        return area

    # ──────────────────────────────────────────────
    #                  右侧信息面板
    # ──────────────────────────────────────────────

    def _draw_panel(self, panel, pw, ph, fps):
        """
        渲染右侧信息面板（带时间感知缓存）。

        面板内容（从上到下）：
            1. 标题区：系统名称 + 当前日期 + 时间（大字）。
            2. 今日统计：已签到/已签退/未打卡/注册总数四项数据块。
            3. 分隔线。
            4. 最近打卡：最新 5 条打卡流水记录。
            5. 导出按钮提示（按 E）。
            6. 当前在场人员列表（视频区当前帧检测到的已知人员）。
            7. FPS 显示 + 退出提示（底部）。

        缓存策略：
            若距上次渲染不超过 _PANEL_TTL 秒且面板尺寸未变，
            直接返回缓存图像，避免每帧都执行 PIL 文字渲染。

        Args:
            panel (np.ndarray): 空白面板图像（BGR，将被填充内容后返回）。
            pw    (int)        : 面板宽度（像素）。
            ph    (int)        : 面板高度（像素）。
            fps   (int)        : 当前帧率（用于底部 FPS 显示）。

        Returns:
            np.ndarray: 渲染完成的面板图像。
        """
        now = time.time()
        # 缓存有效性检查：时间未超期且尺寸一致
        if (self._panel_cache is not None
                and now - self._panel_cache_ts < self._PANEL_TTL
                and self._panel_cache.shape[:2] == (ph, pw)):
            return self._panel_cache

        # 面板背景填充
        panel[:] = C_BG_PANEL

        # ── 标题区背景与分隔线 ──
        cv2.rectangle(panel, (0, 0), (pw, 68), (35, 35, 35), -1)
        cv2.line(panel, (0, 68), (pw, 68), C_ACCENT, 2)

        # ── 获取统计数据 ──
        total, si, so, nc = self._get_stats()
        date_s = datetime.now().strftime("%Y-%m-%d")
        time_s = datetime.now().strftime("%H:%M:%S")

        # ── 统计数据块配置：(标签, 数值, 颜色, y坐标) ──
        stats_cfg = [
            ("已签到",   str(si),    C_GREEN,  204),
            ("已签退",   str(so),    C_CYAN,   248),
            ("未打卡",   str(nc),    C_ORANGE, 292),
            ("注册总数", str(total), C_WHITE,  336),
        ]
        # 统计数据块：半透明背景 + 左侧彩色色条
        for _, _, col, y in stats_cfg:
            blend_rect(panel, (10, y), (pw-10, y+36), (50, 50, 50), 0.5)
            cv2.rectangle(panel, (10, y), (pw-10, y+36), C_BORDER, 1)
            cv2.rectangle(panel, (10, y), (14, y+36), col, -1)

        # ── 最近 5 条打卡记录（按时间倒序） ──
        recs_all = sorted(
            [(n, e["type"], e["time"])
             for n, logs in self.attendance_log.items() for e in logs],
            key=lambda x: x[2], reverse=True
        )[:5]

        # 最近打卡记录块：半透明背景框
        y_rec = 432
        for rn, rt, rtm in recs_all:
            col = C_GREEN if rt == "签到" else C_ORANGE
            blend_rect(panel, (10, y_rec), (pw-10, y_rec+44), (50, 50, 50), 0.4)
            cv2.rectangle(panel, (10, y_rec), (pw-10, y_rec+44), C_BORDER, 1)
            cv2.rectangle(panel, (10, y_rec), (14, y_rec+44), col, -1)
            y_rec += 50

        # ── 导出按钮提示区 ──
        btn_y = y_rec + 8
        blend_rect(panel, (10, btn_y), (pw-10, btn_y+36), (0, 80, 50), 0.55)
        cv2.rectangle(panel, (10, btn_y), (pw-10, btn_y+36), C_ACCENT, 1)

        # ── 当前在场人员（本帧视频区检测到的已知人员） ──
        in_scene = [name for (_, _, _, _, name) in self.last_faces if name]
        scene_y  = btn_y + 44
        if in_scene:
            cv2.line(panel, (12, scene_y), (pw-12, scene_y), C_BORDER, 1)
            scene_y += 8

        # ── 批量收集所有文字绘制指令 ──
        text_items = [
            # 标题区
            ("人脸考勤",     (18, 10),  C_ACCENT,  26, True),
            ("智能管理系统", (14, 38),  C_GRAY,    17, False),
            # 日期与时间
            (date_s,         (10, 82),  C_GRAY,    18, False),
            (time_s,         (6,  106), C_WHITE,   34, True),
            # 统计区标题
            ("▌ 今日统计",   (10, 170), C_ACCENT2, 20, True),
        ]

        # 统计数据块文字（标签 + 数值）
        for label, val, col, y in stats_cfg:
            text_items += [
                (label, (20,     y+6),  C_GRAY, 18, False),
                (val,   (pw-38,  y+4),  col,    23, True),
            ]

        # 最近打卡区标题与分隔线
        cv2.line(panel, (12, 384), (pw-12, 384), C_BORDER, 1)
        text_items.append(("▌ 最近打卡", (10, 396), C_ACCENT2, 20, True))

        # 最近打卡记录文字
        y_rec2 = 432
        for rn, rt, rtm in recs_all:
            col = C_GREEN if rt == "签到" else C_ORANGE
            # 超长姓名截断显示
            dn  = rn[:7] + ".." if len(rn) > 7 else rn
            ts  = rtm[11:]  # 仅显示 HH:MM:SS 部分
            text_items += [
                (dn,  (18,     y_rec2+2),  C_WHITE, 17, False),
                (rt,  (18,     y_rec2+24), col,     15, True),
                (ts,  (pw-86,  y_rec2+14), C_GRAY,  15, False),
            ]
            y_rec2 += 50

        # 导出按钮文字
        text_items.append(("按 E 导出Excel", (16, btn_y+9), C_ACCENT, 17, True))

        # 当前在场人员列表（最多显示 6 人）
        if in_scene:
            text_items.append(("▌ 当前在场", (10, scene_y), C_ACCENT2, 19, True))
            scene_y += 30
            for sname in in_scene[:6]:
                text_items.append((sname, (14, scene_y), C_GREEN, 17, False))
                scene_y += 24

        # 底部信息：FPS + 退出提示
        text_items += [
            (f"FPS: {fps}", (12, ph-52), C_GRAY, 17, False),
            ("按 Q 退出",   (12, ph-30), C_GRAY, 16, False),
        ]

        # 一次批量渲染所有文字
        panel = draw_texts(panel, text_items)

        # 更新缓存
        self._panel_cache    = panel.copy()
        self._panel_cache_ts = now
        return panel

    # ──────────────────────────────────────────────
    #                    主循环
    # ──────────────────────────────────────────────

    def run(self):
        """
        系统主循环：采集视频帧 → 检测识别 → 绘制 UI → 响应键盘事件。

        帧处理流程：
            1. 读取摄像头帧并水平翻转（镜像显示）。
            2. 每 DETECT_INTERVAL 帧执行一次人脸检测与识别。
            3. 用 MultiPersonCounter 统计连续帧数，达到阈值触发打卡。
            4. crop_fill 将摄像头帧适配到视频显示区尺寸。
            5. 在视频区绘制人脸框、通知；在面板区绘制统计信息。
            6. 合并视频区与面板区为最终显示帧并推送到窗口。

        键盘事件：
            Q → 退出循环，导出报表后关闭窗口。
            E → 在后台线程中异步导出 Excel 报表。

        注意：
            - 窗口设置为全屏模式，自动适配分辨率。
            - display_buf 预分配内存，避免每帧 np.hstack 的内存分配开销。
            - 非检测帧复用 last_faces，保持人脸框连续渲染不闪烁。
        """
        cap = cv2.VideoCapture(0)
        cap.set(cv2.CAP_PROP_FRAME_WIDTH,  CAMERA_WIDTH)
        cap.set(cv2.CAP_PROP_FRAME_HEIGHT, CAMERA_HEIGHT)
        cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)  # 最小缓冲区，降低延迟

        # 创建全屏窗口
        cv2.namedWindow(WINDOW_NAME, cv2.WINDOW_NORMAL)
        cv2.setWindowProperty(WINDOW_NAME,
                              cv2.WND_PROP_FULLSCREEN,
                              cv2.WINDOW_FULLSCREEN)

        fps_t = time.time(); fps_cnt = fps = 0
        display_buf = None  # 最终合并帧的预分配缓冲区

        while True:
            ret, frame = cap.read()
            if not ret: break
            frame = cv2.flip(frame, 1)  # 水平翻转，实现镜像效果
            orig_h, orig_w = frame.shape[:2]

            # ── FPS 计算 ──
            fps_cnt += 1
            if time.time() - fps_t >= 1.0:
                fps = fps_cnt; fps_cnt = 0; fps_t = time.time()

            self.frame_count += 1

            # ── 人脸检测与识别（每 DETECT_INTERVAL 帧执行一次）──
            if self.frame_count % DETECT_INTERVAL == 0:
                faces = self.app.get(frame)
                detected_names = set()  # 本帧识别成功的人名集合
                results = []            # 本帧所有人脸的检测结果

                for face in faces:
                    x1, y1, x2, y2 = face.bbox.astype(int)
                    name = self._recognize(face.normed_embedding)
                    if name:
                        detected_names.add(name)
                    results.append((x1, y1, x2, y2, name))

                # 多人帧计数器：返回本帧应触发打卡的人名集合
                to_attend = self._counter.update(detected_names)

                # 对满足条件的每个人执行打卡（支持多人同帧并发）
                for name in to_attend:
                    self._do_attendance(name)

                # 保存检测结果供非检测帧复用（避免帧间闪烁）
                self.last_faces = results

            # ── 获取当前窗口实际尺寸 ──
            try:
                wr = cv2.getWindowImageRect(WINDOW_NAME)
                ww, wh = wr[2], wr[3]
                if ww <= 0 or wh <= 0: raise ValueError
            except Exception:
                ww, wh = 1920, 1080  # 获取失败时使用默认全屏分辨率

            # ── 计算布局尺寸 ──
            pw = max(240, int(ww * PANEL_RATIO))  # 面板宽度（至少 240px）
            vw = ww - pw                           # 视频区宽度
            vh = wh                                # 视频区高度（同窗口高度）

            # ── 缩放参数（用于人脸框坐标变换） ──
            sc = max(vw/orig_w, vh/orig_h)  # crop_fill 使用的缩放比
            sx = sy = sc
            cx = (int(orig_w*sc) - vw) // 2  # X 方向裁剪偏移
            cy = (int(orig_h*sc) - vh) // 2  # Y 方向裁剪偏移

            # ── 渲染各区域 ──
            video = self._crop_fill(frame, vw, vh)          # ① 视频区适配
            video = self._draw_faces(video, sx, sy, cx, cy) # ② 人脸框与标签
            video = self._draw_notifications(video, vw, vh) # ③ 打卡通知
            cv2.line(video, (vw-1, 0), (vw-1, vh), C_ACCENT, 2)  # ④ 分隔线

            panel = np.zeros((vh, pw, 3), dtype=np.uint8)
            panel = self._draw_panel(panel, pw, vh, fps)    # ⑤ 信息面板

            # ── 合并视频区与面板区（预分配缓冲区，避免每帧内存分配）──
            if display_buf is None or display_buf.shape != (vh, ww, 3):
                display_buf = np.empty((vh, ww, 3), dtype=np.uint8)
            display_buf[:, :vw]      = video
            display_buf[:, vw:vw+pw] = panel

            cv2.imshow(WINDOW_NAME, display_buf)

            # ── 键盘事件处理 ──
            key = cv2.waitKey(1) & 0xFF
            if key == ord('q'):
                break                  # 退出主循环
            elif key == ord('e'):
                # 在后台线程中异步导出，不阻塞视频流
                threading.Thread(target=self._do_export, daemon=True).start()

        cap.release()
        cv2.destroyAllWindows()
        print("退出时自动导出报表...")
        self._do_export()  # 退出时同步导出最终报表

    def _do_export(self):
        """
        执行 Excel 报表导出（捕获所有异常，防止导出失败影响主系统）。

        可在主线程（退出时）或后台线程（按 E 时）中调用。
        """
        try:
            export_to_excel(self.attendance_log, list(self.face_db.keys()))
        except Exception as ex:
            print(f"❌ 导出失败: {ex}")


# ============================================================
#                         程序入口
# ============================================================

if __name__ == "__main__":
    system = AttendanceSystem()
    system.run()
